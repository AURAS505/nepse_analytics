# my_portfolio/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, Http404
from django.db import connection, transaction as db_transaction
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Transaction
from listed_companies.models import Companies
# --- RESTORED IMPORT ---
from .utils import calculate_pma_details, calculate_overall_portfolio 

import pandas as pd
import csv
from io import TextIOWrapper, BytesIO
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from collections import defaultdict
from datetime import datetime, date
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
from .models import Transaction, BrokerTransaction
from nepse_data.models import Brokers
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .utils import calculate_pma_details, calculate_overall_portfolio, get_holdings_on_date
from adjustments_stock_price.models import PriceAdjustments


from django.db import connection, transaction as db_transaction # db_transaction is needed for @db_transaction.atomic
from django.http import JsonResponse, HttpResponse, Http404 # HttpResponse is needed for download
import csv
from io import TextIOWrapper, BytesIO # needed for file handling
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime, date # needed for datetime.strptime
from nepse_data.models import Brokers # Brokers model is crucial
from django.db.models import Sum # We'll need this for the opening balance
from django.http import JsonResponse
from django.db.models import Sum, Q, F, DecimalField
from django.db.models.functions import Coalesce
from datetime import date, timedelta
from django.db.models import Min, Max
from nepse_data.models import DividendHistory, StockPrices
from nepse_data.models import CompanyMetrics
try:
    from adjustments_stock_price.models import StockPricesAdj
except ImportError:
    StockPricesAdj = None
from nepse_data.models import Indices
from .models import DematAccount, MeroShareHolding
from .forms import DematAccountForm, MeroShareUploadForm, TransferRequestUploadForm

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .utils import _get_valuation_data
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from openpyxl.worksheet.page import PageMargins


# --- Helper Functions ---


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def fmt_currency_short(value):
    if value is None: return "-"
    try:
        val = Decimal(str(value))
    except:
        return value
    if val == 0: return "-"
    abs_val = abs(val)
    if abs_val >= 10000000: return f"{val/10000000:.2f}C"
    elif abs_val >= 100000: return f"{val/100000:.2f}L"
    elif abs_val >= 1000:   return f"{val/1000:.2f}T"
    else: return f"{val:,.0f}"

# ### --- THIS IS THE FULLY CORRECTED VALUATION FUNCTION --- ###
def _get_valuation_data(start_date, end_date):
    """
    Calculates the full valuation report, including opening,
    movements (buy/sale/bonus), and closing balances.
    """
    
    # 1. Fetch ALL Transactions up to end_date
    transactions = Transaction.objects.filter(
        date__lte=end_date
    ).select_related('symbol').order_by('symbol__sector', 'symbol__script_ticker', 'date', 'created_at')

    # 2. Fetch Latest Prices
    latest_prices = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                    WHERE business_date <= %s
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """, [end_date])
            for row in dictfetchall(cursor):
                latest_prices[row['symbol']] = {
                    'close_price': row.get('close_price') or Decimal('0.0'),
                    'business_date': row.get('business_date')
                }
    except Exception as e:
        print(f"Error fetching prices: {e}")

    # 3. Group transactions by symbol
    grouped_txns = defaultdict(list)
    for txn in transactions:
        grouped_txns[txn.symbol].append(txn)

    sector_grouped_data = defaultdict(list)
    sector_totals = defaultdict(lambda: defaultdict(Decimal))
    grand_totals = defaultdict(lambda: Decimal('0.0'))
    
    TYPE_OPENING = {'Balance b/d'}
    TYPE_SIMPLE_PURCHASE = {'BUY', 'CONVERSION(+)', 'SUSPENSE(+)', 'RIGHT', 'IPO'}
    TYPE_PROPORTIONAL = {'BONUS'} 
    TYPE_SALES = {'SALE', 'CONVERSION(-)', 'SUSPENSE(-)'}
    TYPE_CASH = {'CASH'}

    # 4. Main Logic Loop
    for symbol_obj, txns in grouped_txns.items():
        row = defaultdict(Decimal)
        row.update({
            'company': symbol_obj.script_ticker,
            'company_name': symbol_obj.company_name,
            'sector': symbol_obj.sector,
        })

        global_kitta = 0
        global_cost = Decimal('0.0')
        
        # --- A. Calculate Opening Balance (all txns *before* start_date) ---
        for txn in txns:
            if txn.date < start_date:
                t_type = txn.transaction_type
                kitta = int(txn.kitta or 0)
                amount = txn.billed_amount if txn.billed_amount else Decimal('0.0')
                
                if t_type in TYPE_OPENING or t_type in TYPE_SIMPLE_PURCHASE or t_type in TYPE_PROPORTIONAL:
                    global_kitta += kitta
                    global_cost += amount
                elif t_type in TYPE_SALES:
                    wacc = (global_cost / Decimal(global_kitta)) if global_kitta > 0 else Decimal('0.0')
                    cons = (Decimal(kitta) * wacc).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    global_kitta -= kitta
                    global_cost -= cons
                elif t_type in TYPE_CASH:
                    # NOTE: realized_pl *includes* cash_dividend
                    #row['realized_pl'] += amount 
                    row['cash_dividend'] += amount

        row['op_kitta'] = global_kitta
        row['op_amt'] = global_cost
        row['op_rate'] = (row['op_amt'] / row['op_kitta']) if row['op_kitta'] > 0 else Decimal('0.0')

        period_total_cost = row['op_amt']
        period_total_qty = row['op_kitta']
        period_sales = [] 
        
        # --- B. Calculate Movements (txns *within* date range) ---
        for txn in txns:
            if txn.date >= start_date:
                t_type = txn.transaction_type
                kitta = int(txn.kitta or 0)
                amount = txn.billed_amount if txn.billed_amount else Decimal('0.0')
                rate = txn.eff_rate if txn.eff_rate else Decimal('0.0')

                if t_type in TYPE_OPENING:
                    row['op_kitta'] += kitta; row['op_amt'] += amount
                    period_total_qty += kitta; period_total_cost += amount
                
                elif t_type in TYPE_SIMPLE_PURCHASE:
                    row['buy_kitta'] += kitta; row['buy_amt'] += amount
                    period_total_qty += kitta; period_total_cost += amount
                
                elif t_type in TYPE_PROPORTIONAL:
                    row['bonus_kitta'] += kitta 
                    if amount > 0: row['bonus_amt'] += amount
                    period_total_qty += kitta; period_total_cost += amount
                
                elif t_type in TYPE_SALES:
                    row['sale_kitta'] += kitta; row['sale_amt'] += amount
                    period_sales.append({'kitta': kitta, 'amount': amount, 'rate': rate})
                
                elif t_type in TYPE_CASH:
                    #row['realized_pl'] += amount
                    row['cash_dividend'] += amount

        # --- C. Process Sales ---
        if period_total_qty > 0:
            period_wacc_rate = period_total_cost / Decimal(period_total_qty)
        else:
            period_wacc_rate = Decimal('0.0')

        for sale in period_sales:
            sell_qty = sale['kitta']
            cons = (Decimal(sell_qty) * period_wacc_rate).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            profit = sale['amount'] - cons
            
            row['consumption'] += cons
            row['realized_pl'] += profit
            
            period_total_qty -= sell_qty
            period_total_cost -= cons
            if period_total_qty > 0:
                period_wacc_rate = period_total_cost / Decimal(period_total_qty)
            else:
                period_wacc_rate = Decimal('0.0')

        # --- D. Calculate Closing Balance ---
        row['cl_kitta'] = period_total_qty
        row['cl_cost'] = period_total_cost if period_total_qty > 0 else Decimal('0.0')
        row['cl_rate'] = (row['cl_cost'] / row['cl_kitta']) if row['cl_kitta'] > 0 else Decimal('0.0')

        # --- E. Calculate Rates & Market Value ---
        row['buy_rate'] = (row['buy_amt'] / row['buy_kitta']) if row['buy_kitta'] > 0 else 0
        row['bonus_rate'] = (row['bonus_amt'] / row['bonus_kitta']) if row['bonus_kitta'] > 0 else 0
        row['sale_rate'] = (row['sale_amt'] / row['sale_kitta']) if row['sale_kitta'] > 0 else 0

        price_info = latest_prices.get(symbol_obj.script_ticker, {})
        ltp = price_info.get('close_price', Decimal('0.0'))
        row['ltp'] = ltp
        row['market_val'] = (Decimal(row['cl_kitta']) * ltp).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        row['unrealized_pl'] = row['market_val'] - row['cl_cost']
        row['total_pl'] = row['realized_pl'] + row['unrealized_pl']
        
        # --- FIX: Calculate Realized P/L (ex-cash div) ---
        # This is the value you want to show in the "Performance" column
        row['realized_pl_calc'] = row['realized_pl']
        # This is the value for the final "Total (incl. Div)" column
        row['total_pl_incl_div'] = row['total_pl'] + row['cash_dividend']


        # --- F. Add to totals ---
        if any([row['op_kitta'], row['buy_kitta'], row['bonus_kitta'], row['sale_kitta'], row['cl_kitta'], row['cash_dividend']]):
            sector_grouped_data[row['sector']].append(row)
            st = sector_totals[row['sector']]
            for key, val in row.items():
                if isinstance(val, Decimal):
                    st[key] += val
                elif isinstance(val, (int, float)):
                    st[key] += Decimal(str(val))
            
            for key, val in row.items():
                if isinstance(val, Decimal):
                    grand_totals[key] += val
                elif isinstance(val, (int, float)):
                    grand_totals[key] += Decimal(str(val))

    # --- 6. Format Final Data ---
    sorted_sectors = sorted(sector_grouped_data.keys())
    sn_counter = 1
    final_data = {}
    for sector in sorted_sectors:
        rows = sector_grouped_data[sector]
        rows.sort(key=lambda x: x['company']) 
        for r in rows: r['sn'] = sn_counter; sn_counter += 1
        final_data[sector] = {'rows': rows, 'totals': sector_totals[sector]}
        
    return final_data, grand_totals

def get_fundamental_metrics(symbol_list):
    """
    MASTER FUNCTION: Fetches metrics for BOTH Dashboard and Portfolio Watch.
    - Dashboard needs: EPS, BVP, Mean Returns, Risk-Return, 4W Change, VaR 1W
    - Watch needs: VaR 1W, VaR 1M, Beta
    """
    if not symbol_list:
        return {}

    metrics_map = {}
    
    # --- 1. Fetch ALL Metrics from nepse_data.company_metrics ---
    try:
        with connection.cursor() as cursor:
            placeholders = ','.join(['%s'] * len(symbol_list))
            
            # Combined Query: Includes items for Dashboard AND Portfolio Watch
            query = f"""
                WITH LatestMetrics AS (
                    SELECT 
                        symbol, 
                        metric_item, 
                        metric_value,
                        ROW_NUMBER() OVER(PARTITION BY symbol, metric_item ORDER BY business_date DESC) as rn
                    FROM company_metrics 
                    WHERE symbol IN ({placeholders})
                    AND metric_item IN (
                        'VaR 1 Week @5%%', 
                        'VaR 1 Month @5%%',             -- Needed for Watch
                        'Mean returns [w,1Y]%%',        -- Needed for Dashboard
                        'Estimated risk-return trade-off [w,1Y]', -- Needed for Dashboard
                        'Beta Weekly',                  -- Needed for Watch & Dashboard
                        'EPS (D)',                      -- Needed for Dashboard (PE calc)
                        'Bvp'                           -- Needed for Dashboard
                    )
                )
                SELECT symbol, metric_item, metric_value 
                FROM LatestMetrics 
                WHERE rn = 1;
            """
            
            cursor.execute(query, symbol_list)
            rows = dictfetchall(cursor)
            
            for row in rows:
                sym = row['symbol']
                if sym not in metrics_map: metrics_map[sym] = {}
                
                item = row['metric_item']
                val = row['metric_value']
                
                # Map to keys used by BOTH views
                if item == 'VaR 1 Week @5%': 
                    metrics_map[sym]['var_1w'] = val
                elif item == 'VaR 1 Month @5%': 
                    metrics_map[sym]['var_1m'] = val
                elif item == 'Beta Weekly': 
                    metrics_map[sym]['beta'] = val
                elif item == 'Mean returns [w,1Y]%': 
                    metrics_map[sym]['mean_returns'] = val
                elif item == 'Estimated risk-return trade-off [w,1Y]': 
                    metrics_map[sym]['risk_return'] = val
                elif item == 'EPS (D)': 
                    metrics_map[sym]['eps'] = val
                elif item == 'Bvp': 
                    metrics_map[sym]['bvp'] = val

    except Exception as e:
        print(f"Error fetching company metrics: {e}")

    # --- 2. Calculate 4-Week Price Change % (CRITICAL FOR DASHBOARD) ---
    # Use StockPricesAdj if available, else StockPrices
    PriceModel = StockPricesAdj if StockPricesAdj else StockPrices
    price_field = 'close_price_adj' if StockPricesAdj else 'close_price'

    today = date.today()
    four_weeks_ago = today - timedelta(days=28)
    
    for sym in symbol_list:
        if sym not in metrics_map: metrics_map[sym] = {}
        
        try:
            # Get Latest Price
            latest_obj = PriceModel.objects.filter(symbol=sym).order_by('-business_date').first()
            # Get Old Price (closest to 4 weeks ago)
            old_obj = PriceModel.objects.filter(symbol=sym, business_date__lte=four_weeks_ago).order_by('-business_date').first()

            current_p = getattr(latest_obj, price_field, 0) if latest_obj else 0
            old_p = getattr(old_obj, price_field, 0) if old_obj else 0

            if current_p and old_p and old_p > 0:
                metrics_map[sym]['four_week_change'] = ((current_p - old_p) / old_p) * 100
            else:
                metrics_map[sym]['four_week_change'] = 0
                
        except Exception as e:
            metrics_map[sym]['four_week_change'] = 0

    return metrics_map

# ### --- END OF FIXED FUNCTION --- ###
# --- STANDARD VIEWS ---
@login_required
def portfolio_home(request):
    stats = {
        'total_scrips_traded': 0,
        'total_holdings': 0,
        'available_shares': 0,
        'total_investment': Decimal('0.0'),
        'total_market_value': Decimal('0.0'),
        'total_profit_loss': Decimal('0.0'),
        'realized_pl': Decimal('0.0'),
        'unrealized_pl': Decimal('0.0'),
        'total_investment_crore': Decimal('0.0'),
        'total_market_value_crore': Decimal('0.0'),
        'total_profit_loss_crore': Decimal('0.0'),
        'top_investments': [],
        'top_gainers': [],
        'top_losers': []
    }
    
    # 1. Fetch Market Cap Stats
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT total_scrips_traded FROM marcap ORDER BY business_date DESC LIMIT 1")
            result = cursor.fetchone()
            stats['total_scrips_traded'] = result[0] if result and result[0] is not None else 0
    except Exception as e:
        print(f"Error fetching marcap: {e}")

    # 2. Fetch Latest Prices & Portfolio Data
    try:
        latest_prices = {}
        with connection.cursor() as cursor:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """)
            for row in dictfetchall(cursor):
                latest_prices[row['symbol']] = {
                    'close_price': row.get('close_price') or Decimal('0.0'),
                    'business_date': row.get('business_date')
                }

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT SUM(kitta),
                    SUM(CASE WHEN transaction_type IN ('Balance b/d', 'BUY', 'IPO', 'RIGHT', 'CONVERSION(+)', 'SUSPENSE(+)', 'CASH') THEN billed_amount
                        WHEN transaction_type IN ('SALE', 'CONVERSION(-)', 'SUSPENSE(-)') THEN -billed_amount ELSE 0 END)
                FROM my_portfolio_transaction
            """)
            summary_row = cursor.fetchone()
            stats['total_holdings'] = summary_row[0] or 0
            stats['total_investment'] = summary_row[1] or Decimal('0.0')

            cursor.execute("SELECT * FROM my_portfolio_transaction ORDER BY symbol_id, date, created_at")
            all_transactions = dictfetchall(cursor)

        # 3. Calculate Basic Portfolio Stats
        overall_stats, holdings_summary_list = calculate_overall_portfolio(all_transactions, latest_prices)
        
        # --- NEW STEP: Fetch Fundamental Metrics (VaR, Beta, etc.) ---
        # Extract list of symbols we currently hold
        held_symbols = [h['symbol'] for h in holdings_summary_list]
        # Call our helper function
        fundamental_data = get_fundamental_metrics(held_symbols)
        
        sector_book_values = defaultdict(Decimal)
        portfolio_book_value = overall_stats.get('book_value', Decimal('0.0'))
        for h in holdings_summary_list:
            sector = h.get('sector', 'Unknown')
            sector_book_values[sector] += h['book_value']

        # 4. Enrich Holdings with both Portfolio Stats and Fundamental Metrics
        enriched_holdings = []
        for h in holdings_summary_list:
            # Standard Portfolio Stats
            book_val = h['book_value']
            total_pl = h['realized_pl'] + h['unrealized_pl']
            sector = h.get('sector', 'Unknown')
            sec_book_val = sector_book_values[sector]
            
            h['allocation_sector'] = (book_val / sec_book_val * 100) if sec_book_val > 0 else Decimal(0)
            h['allocation_total'] = (book_val / portfolio_book_value * 100) if portfolio_book_value > 0 else Decimal(0)
            h['roi_individual'] = (total_pl / book_val * 100) if book_val > 0 else Decimal(0)
            h['contribution_sector'] = (total_pl / sec_book_val * 100) if sec_book_val > 0 else Decimal(0)
            h['contribution_total'] = (total_pl / portfolio_book_value * 100) if portfolio_book_value > 0 else Decimal(0)
            h['total_pl'] = total_pl
            
            # --- NEW: Attach Fundamental Metrics ---
            sym = h['symbol']
            metrics = fundamental_data.get(sym, {})
            
            h['var_1w'] = metrics.get('var_1w')
            h['mean_returns'] = metrics.get('mean_returns')
            h['risk_return'] = metrics.get('risk_return')
            h['beta'] = metrics.get('beta')
            h['bvp'] = metrics.get('bvp')
            h['four_week_change'] = metrics.get('four_week_change', 0)
            
            # EPS and PE Calculation
            h['eps'] = metrics.get('eps')
            
            # Calculate PE Ratio: Market Price / EPS
            # We use the LTP from latest_prices which is already in 'h'
            if h['ltp'] and h['eps'] and float(h['eps']) != 0:
                try:
                    h['pe_ratio'] = float(h['ltp']) / float(h['eps'])
                except:
                    h['pe_ratio'] = 0
            else:
                h['pe_ratio'] = 0

            enriched_holdings.append(h)

        # 5. Finalize Stats for Template
        CRORE = Decimal('10000000.0')
        if stats['total_investment']: stats['total_investment_crore'] = stats['total_investment'] / CRORE
        
        stats['total_market_value'] = overall_stats.get('market_value', Decimal('0.0'))
        if stats['total_market_value']: stats['total_market_value_crore'] = stats['total_market_value'] / CRORE
        
        stats['total_profit_loss'] = overall_stats.get('total_profit', Decimal('0.0'))
        if stats['total_profit_loss']: stats['total_profit_loss_crore'] = stats['total_profit_loss'] / CRORE
        
        stats['realized_pl'] = overall_stats.get('realized_pl', Decimal('0.0'))
        stats['unrealized_pl'] = overall_stats.get('unrealized_pl', Decimal('0.0'))
        
        stats['available_shares'] = sum(h['closing_kitta'] for h in holdings_summary_list)
        stats['holdings_count'] = len(holdings_summary_list)
        
        # Sort top investments by Book Value (Investment amount)
        stats['top_investments'] = sorted(enriched_holdings, key=lambda x: x['book_value'], reverse=True)[:10]
        
        # Gainers/Losers
        gainers = [h for h in enriched_holdings if h['total_pl'] >= 0]
        stats['top_gainers'] = sorted(gainers, key=lambda x: x['total_pl'], reverse=True)[:5]
        
        losers = [h for h in enriched_holdings if h['total_pl'] < 0]
        stats['top_losers'] = sorted(losers, key=lambda x: x['total_pl'])[:10]

    except Exception as e:
        messages.error(request, f"Could not load portfolio statistics: {e}")
    
    return render(request, 'my_portfolio/dashboard.html', {'stats': stats})



@login_required
def transaction_list_and_add(request):
    # (This view is correct and remains unchanged)
    if request.method == 'POST':
        try:
            def to_decimal_or_none(value_str):
                if value_str: return Decimal(value_str)
                return None
            def to_int_or_none(value_str):
                if value_str: return int(value_str)
                return None

            data = request.POST
            date = data.get('date')
            symbol_ticker = data.get('symbol', '').upper()
            transaction_type = data.get('transaction_type')
            
            if not date or not symbol_ticker or not transaction_type:
                return JsonResponse({"message": "Error: Missing required fields (Date, Symbol, Type)."}, status=400)
            
            try:
                company = Companies.objects.get(script_ticker=symbol_ticker)
            except Companies.DoesNotExist:
                return JsonResponse({"message": f"Invalid symbol. Company '{symbol_ticker}' not found in database."}, status=400)

            new_txn = Transaction(
                date=date,
                symbol=company,
                transaction_type=transaction_type,
                kitta = to_int_or_none(data.get('kitta')),
                broker = data.get('broker') or None,
                rate = to_decimal_or_none(data.get('rate')),
                gross_amount = to_decimal_or_none(data.get('gross_amount')),
                commission_rate = to_decimal_or_none(data.get('commission_rate')),
                commission_amount = to_decimal_or_none(data.get('commission_amount')),
                nepse_commission = to_decimal_or_none(data.get('nepse_commission')),
                sebon_regularity_fee = to_decimal_or_none(data.get('sebon_regularity_fee')),
                broker_commission = to_decimal_or_none(data.get('broker_commission')),
                sebo_commission = to_decimal_or_none(data.get('sebo_commission')),
                cgt = to_decimal_or_none(data.get('cgt')),
                dp_fee = to_decimal_or_none(data.get('dp_fee')),
                billed_amount = to_decimal_or_none(data.get('billed_amount'))
            )
            
            new_txn.save()
            
            return JsonResponse({"message": "Transaction added successfully!", "unique_id": new_txn.unique_id}, status=200)
        
        except InvalidOperation:
            return JsonResponse({"message": "Error: Invalid number format for a financial field."}, status=400)
        except Exception as e:
            return JsonResponse({"message": f"An unexpected server error occurred: {str(e)}"}, status=500)

    transactions = Transaction.objects.all().select_related('symbol').order_by('-date', '-created_at')
    companies = Companies.objects.all().order_by('script_ticker')
    context = {
        'transactions': transactions, 
        'companies': companies, 
        'transaction_choices': Transaction.TransactionType.choices
    }
    return render(request, 'my_portfolio/transactions.html', context)


# ### MODIFIED VIEW: transaction_list_and_add ###
@login_required
def transaction_list_and_add(request):
    
    # --- NEW POST LOGIC ---
    if request.method == 'POST':
        try:
            # --- Helper to convert form strings to Decimal or None ---
            def to_decimal_or_none(value_str):
                if value_str:
                    return Decimal(value_str)
                return None
            
            # --- Helper to convert form strings to Int or None ---
            def to_int_or_none(value_str):
                if value_str:
                    return int(value_str)
                return None

            # --- 1. Get All 17 Fields from FormData ---
            data = request.POST
            date = data.get('date')
            symbol_ticker = data.get('symbol', '').upper()
            transaction_type = data.get('transaction_type')
            
            if not date or not symbol_ticker or not transaction_type:
                return JsonResponse({"message": "Error: Missing required fields (Date, Symbol, Type)."}, status=400)
            
            try:
                company = Companies.objects.get(script_ticker=symbol_ticker)
            except Companies.DoesNotExist:
                return JsonResponse({"message": f"Invalid symbol. Company '{symbol_ticker}' not found in database."}, status=400)

            # --- 2. Create the Transaction Instance ---
            new_txn = Transaction(
                date=date,
                symbol=company,
                transaction_type=transaction_type,
                
                # Get all other fields, converting to correct type or None
                kitta = to_int_or_none(data.get('kitta')),
                broker = data.get('broker') or None,
                rate = to_decimal_or_none(data.get('rate')),
                gross_amount = to_decimal_or_none(data.get('gross_amount')),
                commission_rate = to_decimal_or_none(data.get('commission_rate')),
                commission_amount = to_decimal_or_none(data.get('commission_amount')),
                nepse_commission = to_decimal_or_none(data.get('nepse_commission')),
                sebon_regularity_fee = to_decimal_or_none(data.get('sebon_regularity_fee')),
                broker_commission = to_decimal_or_none(data.get('broker_commission')),
                sebo_commission = to_decimal_or_none(data.get('sebo_commission')),
                cgt = to_decimal_or_none(data.get('cgt')),
                dp_fee = to_decimal_or_none(data.get('dp_fee')),
                billed_amount = to_decimal_or_none(data.get('billed_amount'))
            )
            
            # 3. Save the object
            # This will also run your model's .save() method,
            # which auto-calculates eff_rate, gross_amount, etc.
            new_txn.save()
            
            return JsonResponse({"message": "Transaction added successfully!", "unique_id": new_txn.unique_id}, status=200)
        
        except InvalidOperation:
            return JsonResponse({"message": "Error: Invalid number format for a financial field."}, status=400)
        except Exception as e:
            return JsonResponse({"message": f"An unexpected server error occurred: {str(e)}"}, status=500)

    # --- GET request logic (unchanged) ---
    transactions = Transaction.objects.all().select_related('symbol').order_by('-date', '-created_at')
    companies = Companies.objects.all().order_by('script_ticker')
    context = {
        'transactions': transactions, 
        'companies': companies, 
        'transaction_choices': Transaction.TransactionType.choices
    }
    return render(request, 'my_portfolio/transactions.html', context)


# ### MODIFIED VIEW: transaction_edit ###
@login_required
def transaction_edit(request, unique_id):
    # (This view is correct and remains unchanged)
    txn = get_object_or_404(Transaction, unique_id=unique_id)
    
    if request.method == 'POST':
        try:
            def to_decimal_or_none(value_str):
                if value_str: return Decimal(value_str)
                return None
            def to_int_or_none(value_str):
                if value_str: return int(value_str)
                return None
            
            data = request.POST
            
            symbol_ticker = data.get('symbol').upper()
            try:
                company = Companies.objects.get(script_ticker=symbol_ticker)
            except Companies.DoesNotExist:
                messages.error(request, "Invalid symbol. Company not found.")
                return redirect('my_portfolio:transaction_edit', unique_id=unique_id)

            txn.date = data.get('date')
            txn.symbol = company
            txn.transaction_type = data.get('transaction_type')
            txn.kitta = to_int_or_none(data.get('kitta'))
            txn.broker = data.get('broker') or None
            txn.rate = to_decimal_or_none(data.get('rate'))
            txn.gross_amount = to_decimal_or_none(data.get('gross_amount'))
            txn.commission_rate = to_decimal_or_none(data.get('commission_rate'))
            txn.commission_amount = to_decimal_or_none(data.get('commission_amount'))
            txn.nepse_commission = to_decimal_or_none(data.get('nepse_commission'))
            txn.sebon_regularity_fee = to_decimal_or_none(data.get('sebon_regularity_fee'))
            txn.broker_commission = to_decimal_or_none(data.get('broker_commission'))
            txn.sebo_commission = to_decimal_or_none(data.get('sebo_commission'))
            txn.cgt = to_decimal_or_none(data.get('cgt'))
            txn.dp_fee = to_decimal_or_none(data.get('dp_fee'))
            txn.billed_amount = to_decimal_or_none(data.get('billed_amount'))

            txn.save()
            
            messages.success(request, "Transaction updated successfully.")
            return redirect('my_portfolio:transactions')
            
        except Exception as e:
            messages.error(request, f"Error updating transaction: {e}")
            
    companies = Companies.objects.all().order_by('script_ticker')
    context = {
        'transaction': txn, 
        'companies': companies
    }
    return render(request, 'my_portfolio/edit_transaction.html', context)


@login_required
@require_POST
def transaction_delete(request, unique_id):
    # (This view is correct and remains unchanged)
    txn = get_object_or_404(Transaction, unique_id=unique_id)
    try:
        txn.delete()
        messages.success(request, "Transaction deleted.")
    except Exception as e:
        messages.error(request, f"Error deleting transaction: {e}")
    return redirect('my_portfolio:transactions')

@login_required
@require_POST
def transaction_delete_all(request):
    # (This view is correct and remains unchanged)
    try:
        Transaction.objects.all().delete()
        messages.success(request, "All transactions have been deleted.")
    except Exception as e:
        messages.error(request, f"Error deleting all transactions: {e}")
    return redirect('my_portfolio:transactions')

# ### MODIFIED VIEW: transaction_upload ###
@login_required
@require_POST
@db_transaction.atomic
def transaction_upload(request):
    # (This view is correct and remains unchanged)
    file = request.FILES.get('file')
    if not file:
        messages.error(request, "No file selected.")
        return redirect('my_portfolio:transactions')
        
    filename = file.name
    success_count = 0
    error_count = 0
    errors = []

    def to_decimal_or_none(value_str):
        if value_str: return Decimal(str(value_str).strip())
        return None
    def to_int_or_none(value_str):
        if value_str:
            try:
                # 1. Convert to Decimal safely
                dec_val = Decimal(str(value_str).strip())
                # 2. Convert to integer. Use quantize/round if intermediate decimals exist.
                # For Kitta, we assume it's a whole number.
                return int(dec_val.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
            except InvalidOperation:
                # If it's a non-numeric string, int() will fail later, so return None
                return None
        return None

    try:
        required_headers = ['Date', 'Symbol', 'Transaction Type']
        
        if filename.endswith('.csv'):
            csv_file = TextIOWrapper(file, encoding='utf-8', errors='replace')
            reader = csv.DictReader(csv_file)
            reader.fieldnames = [header.strip() for header in reader.fieldnames]
            data_iter = enumerate(reader, start=2)
            headers = reader.fieldnames
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file, dtype=str).fillna('') 
            df.columns = [col.strip() for col in df.columns]
            data_iter = df.iterrows() 
            headers = df.columns
        else:
            messages.error(request, "Unsupported file type. Please upload a CSV or XLSX file.")
            return redirect('my_portfolio:transactions')

        if not all(header in headers for header in required_headers):
            missing_headers = [h for h in required_headers if h not in headers]
            messages.error(request, f"File missing required columns. Missing: {', '.join(missing_headers)}")
            return redirect('my_portfolio:transactions')

        companies_cache = {c.script_ticker: c for c in Companies.objects.all()}
        valid_types = set(Transaction.TransactionType.values)

        for index, row in data_iter:
            row_num = index + 2 if filename.endswith(('.xlsx', '.xls')) else index 
            
            try:
                date_str = str(row.get('Date', '')).split()[0].strip()
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                symbol = str(row.get('Symbol', '')).strip().upper()
                if symbol not in companies_cache: 
                    raise ValueError(f"Symbol '{symbol}' not found")
                company = companies_cache[symbol]
                
                transaction_type = str(row.get('Transaction Type', '')).strip()
                if transaction_type.lower() == 'bonus': transaction_type = 'BONUS'
                elif transaction_type.lower() == 'buy': transaction_type = 'BUY'
                elif transaction_type.lower() == 'sale': transaction_type = 'SALE'
                elif transaction_type.lower() == 'ipo': transaction_type = 'IPO'
                elif transaction_type.lower() == 'right': transaction_type = 'RIGHT'
                elif transaction_type.lower() == 'cash': transaction_type = 'CASH'
                
                if transaction_type not in valid_types: 
                    raise ValueError(f"Invalid Transaction Type '{transaction_type}'")
                
                Transaction(
                    date=date, 
                    symbol=company, 
                    transaction_type=transaction_type, 
                    kitta = to_int_or_none(row.get('Kitta')),
                    broker = str(row.get('Broker', '')).strip() or None,
                    rate = to_decimal_or_none(row.get('rate')),
                    gross_amount = to_decimal_or_none(row.get('gross_amount')),
                    commission_rate = to_decimal_or_none(row.get('commission_rate')),
                    commission_amount = to_decimal_or_none(row.get('commission_amount')),
                    nepse_commission = to_decimal_or_none(row.get('nepse_commission')),
                    sebon_regularity_fee = to_decimal_or_none(row.get('sebon_regularity_fee')),
                    broker_commission = to_decimal_or_none(row.get('broker_commission')),
                    sebo_commission = to_decimal_or_none(row.get('sebo_commission')),
                    cgt = to_decimal_or_none(row.get('cgt')),
                    dp_fee = to_decimal_or_none(row.get('dp_fee')),
                    billed_amount = to_decimal_or_none(row.get('billed_amount'))
                ).save()
                
                success_count += 1
                
            except Exception as e:
                 errors.append(f"Row {row_num}: Error - {str(e)}")
                 error_count += 1
                 continue
                 
        if error_count > 0:
            db_transaction.set_rollback(True)
            messages.error(request, f"Upload failed. {error_count} errors. First error: {errors[0]}")
        else:
            messages.success(request, f"Upload successful! {success_count} transactions added.")
            
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")
        
    return redirect('my_portfolio:transactions')


@login_required
def download_transaction_template(request, file_type):
    # (This view is correct and remains unchanged)
    fieldnames = [
        'Date', 'Symbol', 'Transaction Type', 'Kitta', 'Broker', 
        'rate', 'gross_amount', 'commission_rate', 'commission_amount', 
        'nepse_commission', 'sebon_regularity_fee', 'broker_commission',
        'sebo_commission', 'cgt', 'dp_fee', 'billed_amount'
    ]
    
    sample_data = [
        {
            'Date': '2025-07-17', 'Symbol': 'HBL', 'Transaction Type': 'BUY', 'Kitta': 11050, 'Broker': '35',
            'rate': '249.85', 'gross_amount': '2760842.50', 'commission_rate': '0.004', 
            'commission_amount': '11043.37', 'nepse_commission': '552.17', 'sebon_regularity_fee': '414.13', 
            'broker_commission': None, 'sebo_commission': None, 'cgt': '0', 
            'dp_fee': '25', 'billed_amount': '2772877.17'
        },
        {
            'Date': '2025-07-17', 'Symbol': 'PRIN', 'Transaction Type': 'BONUS', 'Kitta': 4319, 'Broker': '',
            'rate': '0', 'billed_amount': '0'
        },
        {
            'Date': '2025-07-18', 'Symbol': 'NABIL', 'Transaction Type': 'CASH', 'Kitta': None, 'Broker': '',
            'rate': None, 'billed_amount': '15000'
        }
    ]
    
    if file_type == 'csv':
        output = TextIOWrapper(BytesIO(), encoding='utf-8', newline='')
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(sample_data)
        output.flush()
        response = HttpResponse(output.buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transaction_template_detailed.csv"'
        return response
        
    elif file_type == 'excel':
        output = BytesIO()
        df = pd.DataFrame(sample_data, columns=fieldnames)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transactions')
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="transaction_template_detailed.xlsx"'
        return response
        
    return Http404("Invalid file type")


# --- NEW VIEW FUNCTION ---
@login_required
@require_POST
@db_transaction.atomic
def sync_dividend_transactions(request):
    """
    Finds dividends for held stocks and auto-creates
    CASH, BONUS, or RIGHT transactions if they don't already exist.
    """
    try:
        # 1. Get all companies the user has ever held
        held_symbols_tickers = Transaction.objects.values_list('symbol__script_ticker', flat=True).distinct()
        
        # 2. Get all companies (symbols) from the ticker list
        held_companies = Companies.objects.filter(script_ticker__in=held_symbols_tickers)
        
        # 3. Find all dividends for these companies that have a book closure date
        dividends = DividendHistory.objects.filter(
            symbol__in=held_symbols_tickers,
            book_closure_date__isnull=False
        ).order_by('book_closure_date')

        # 4. Create a lookup for companies
        company_lookup = {c.script_ticker: c for c in held_companies}

        created_count = 0
        skipped_count = 0
        
        # 5. Loop through each dividend
        for div in dividends:
            if div.book_closure_date is None:
                continue

            # 6. Get the company object
            company = company_lookup.get(div.symbol)
            if not company:
                continue
                
            # 7. Find user's holdings *before* the book closure date
            holdings = get_holdings_on_date(company, div.book_closure_date)
            
            if holdings > 0:
                # --- A. Check for BONUS ---
                bonus_pct = div.bonus_percent or Decimal('0.0')
                if bonus_pct > 0:
                    bonus_kitta = (Decimal(holdings) * bonus_pct / Decimal('100.0')).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                    # Check for duplicate
                    exists = Transaction.objects.filter(
                        symbol=company,
                        date=div.book_closure_date,
                        transaction_type='BONUS'
                    ).exists()
                    
                    if not exists and bonus_kitta > 0:
                        Transaction.objects.create(
                            symbol=company,
                            date=div.book_closure_date,
                            transaction_type='BONUS',
                            kitta=int(bonus_kitta),
                            billed_amount=Decimal('0.00'),
                            rate=Decimal('0.00')
                        )
                        created_count += 1
                    elif exists:
                        skipped_count += 1

                # --- B. Check for CASH ---
                cash_pct = div.cash_percent or Decimal('0.0')
                if cash_pct > 0:
                    # Cash dividend is (kitta * par (100) * percent)
                    # Or, more simply: kitta * cash_pct (as 10% cash = 10 Rs)
                    cash_amount = (Decimal(holdings) * cash_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                    
                    # Check for duplicate
                    exists = Transaction.objects.filter(
                        symbol=company,
                        date=div.book_closure_date,
                        transaction_type='CASH'
                    ).exists()
                    
                    if not exists and cash_amount > 0:
                        Transaction.objects.create(
                            symbol=company,
                            date=div.book_closure_date,
                            transaction_type='CASH',
                            kitta=None, # No kitta for cash dividend
                            billed_amount=cash_amount
                        )
                        created_count += 1
                    elif exists:
                        skipped_count += 1
                        
                # --- C. Check for RIGHT ---
                right_pct = div.right_percent or Decimal('0.0')
                if right_pct > 0:
                    right_kitta = (Decimal(holdings) * right_pct / Decimal('100.0')).quantize(Decimal('0'), rounding=ROUND_HALF_UP)
                    
                    # Check for duplicate
                    exists = Transaction.objects.filter(
                        symbol=company,
                        date=div.book_closure_date,
                        transaction_type='RIGHT'
                    ).exists()
                    
                    if not exists and right_kitta > 0:
                        # Assume rights are bought at 100 par
                        right_cost = right_kitta * 100
                        Transaction.objects.create(
                            symbol=company,
                            date=div.book_closure_date,
                            transaction_type='RIGHT',
                            kitta=int(right_kitta),
                            billed_amount=right_cost,
                            rate=Decimal('100.00')
                        )
                        created_count += 1
                    elif exists:
                        skipped_count += 1

        return JsonResponse({
            "message": f"Sync complete! Created {created_count} new transactions. Skipped {skipped_count} duplicates."
        }, status=200)

    except Exception as e:
        return JsonResponse({"message": f"An unexpected error occurred: {str(e)}"}, status=500)
# --- END NEW VIEW ---

@login_required
def company_dashboard(request):
    # (This view is correct and remains unchanged)
    latest_prices = {}
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """)
            for row in dictfetchall(cursor):
                latest_prices[row['symbol']] = {
                    'close_price': row.get('close_price') or Decimal('0.0'),
                    'business_date': row.get('business_date')
                }
        except Exception as e:
            print(f"Error fetching latest prices: {e}")
    
    overall_stats, holdings_summary_list = {}, []
    all_transactions = [] 

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM my_portfolio_transaction ORDER BY symbol_id, date, created_at")
            all_transactions = dictfetchall(cursor) 
        
        overall_stats, holdings_summary_list = calculate_overall_portfolio(all_transactions, latest_prices)
    except Exception as e:
        messages.error(request, f"Could not calculate portfolio stats: {e}")
    
    symbol = request.GET.get('symbol')
    company_info, detailed_calculations, summary_data = None, [], None
    if symbol:
        try:
            symbol_txns = [txn for txn in all_transactions if txn['symbol_id'] == symbol]
            
            if symbol_txns:
                company_info = {'symbol': symbol, 'script': symbol_txns[0]['script'], 'sector': symbol_txns[0]['sector']}
                price_info = latest_prices.get(symbol, {})
                detailed_calculations, summary_data = calculate_pma_details(symbol_txns, price_info)
        except Exception as e:
             messages.error(request, f"Could not generate report for {symbol}: {e}")
    
    context = {
        'holdings_list': holdings_summary_list,
        'overall_stats': overall_stats,
        'company': company_info, 
        'details': detailed_calculations, 
        'summary': summary_data,
        'current_symbol': symbol
    }
    return render(request, 'my_portfolio/company_dashboard.html', context)


@login_required
def api_company_details(request, symbol):
    # (This view is correct and remains unchanged)
    try:
        company = Companies.objects.get(script_ticker__iexact=symbol)
        return JsonResponse({'script_ticker': company.script_ticker, 'company_name': company.company_name, 'sector': company.sector})
    except Companies.DoesNotExist:
        return JsonResponse({"error": "Company not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
def valuation_report(request):
    # (This view is correct and remains unchanged)
    end_date_str = request.GET.get('end_date')
    start_date_str = request.GET.get('start_date')
    if end_date_str: end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else: 
        latest_price = StockPrices.objects.order_by('-business_date').first()
        end_date = latest_price.business_date if latest_price else timezone.now().date()
    if start_date_str: start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        first_txn = Transaction.objects.order_by('date').first()
        start_date = first_txn.date if first_txn else date(2000, 1, 1)

    raw_data, raw_grand_totals = _get_valuation_data(start_date, end_date)

    formatted_data = {}
    for sector, content in raw_data.items():
        new_rows = []
        for r in content['rows']:
            nr = r.copy()
            nr['op_amt'] = fmt_currency_short(r['op_amt'])
            nr['buy_amt'] = fmt_currency_short(r['buy_amt'])
            nr['bonus_amt'] = fmt_currency_short(r['bonus_amt'])
            nr['sale_amt'] = fmt_currency_short(r['sale_amt'])
            nr['consumption'] = fmt_currency_short(r['consumption'])
            nr['realized_pl_calc'] = fmt_currency_short(r['realized_pl_calc']) # <-- ADD THIS
            nr['cash_dividend'] = fmt_currency_short(r['cash_dividend'])
            nr['cl_cost'] = fmt_currency_short(r['cl_cost'])
            nr['market_val'] = fmt_currency_short(r['market_val'])
            nr['unrealized_pl'] = fmt_currency_short(r['unrealized_pl'])
            nr['total_pl'] = fmt_currency_short(r['total_pl'])
            nr['total_pl_incl_div'] = fmt_currency_short(r['total_pl_incl_div']) # <-- ADD THIS
            new_rows.append(nr)
        
        t = content['totals']
        new_totals = t.copy()
        new_totals['op_amt'] = fmt_currency_short(t['op_amt'])
        new_totals['buy_amt'] = fmt_currency_short(t['buy_amt'])
        new_totals['bonus_amt'] = fmt_currency_short(t['bonus_amt'])
        new_totals['sale_amt'] = fmt_currency_short(t['sale_amt'])
        new_totals['consumption'] = fmt_currency_short(t['consumption'])
        new_totals['realized_pl_calc'] = fmt_currency_short(t['realized_pl_calc']) # <-- ADD THIS
        new_totals['cash_dividend'] = fmt_currency_short(t['cash_dividend'])
        new_totals['cl_cost'] = fmt_currency_short(t['cl_cost'])
        new_totals['market_val'] = fmt_currency_short(t['market_val'])
        new_totals['unrealized_pl'] = fmt_currency_short(t['unrealized_pl'])
        new_totals['total_pl'] = fmt_currency_short(t['total_pl'])
        new_totals['total_pl_incl_div'] = fmt_currency_short(t['total_pl_incl_div']) # <-- ADD THIS
        
        formatted_data[sector] = {'rows': new_rows, 'totals': new_totals}

    formatted_grand_totals = {}
    for k, v in raw_grand_totals.items():
        formatted_grand_totals[k] = fmt_currency_short(v)

    context = {
        'valuation_data': formatted_data,
        'start_date': start_date,
        'end_date': end_date,
        'grand_totals': formatted_grand_totals,
    }
    return render(request, 'my_portfolio/valuation_report.html', context)


@login_required
def download_valuation_report(request):
    # 1. Date Logic
    end_date_str = request.GET.get('end_date')
    start_date_str = request.GET.get('start_date')
    if end_date_str: 
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else: 
        latest_price = StockPrices.objects.order_by('-business_date').first()
        end_date = latest_price.business_date if latest_price else timezone.now().date()
    if start_date_str: 
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        first_txn = Transaction.objects.order_by('date').first()
        start_date = first_txn.date if first_txn else date(2000, 1, 1)

    # 2. Get Data
    data, totals = _get_valuation_data(start_date, end_date)

    # 3. Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Report"
    
    # --- GLOBAL SETTINGS ---
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.sheet_properties.outlinePr.summaryBelow = False
    
    # --- STYLES ---
    font_header = Font(name='Calibri', size=9, bold=True)
    font_body = Font(name='Calibri', size=9)
    font_subtotal = Font(name='Calibri', size=9, bold=True)
    font_grand = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    fill_header = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_subtotal = PatternFill(start_color="DFE1E5", end_color="DFE1E5", fill_type="solid") 
    fill_grand = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_profit = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    fill_loss = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    thin_side = Side(style='thin', color="E2E2E2")
    thick_side = Side(style='medium', color="999999")
    
    num_fmt = '#,##0'
    dec_fmt = '#,##0.00'

    # --- HEADERS ---
    ltp_header = f"LTP\n{end_date.strftime('%Y-%m-%d')}"
    
    headers_cat = [
        ("S.N.", 1), ("Symbol", 1), ("Company Name", 1),
        ("Opening", 3), ("Purchase", 3), ("Bonus", 3), ("Sales", 3), 
        ("Performance", 2), 
        ("Closing (Cost)", 3), ("Market Valuation", 3), 
        ("Net P/L", 3)
    ]
    
    col = 1
    for title, span in headers_cat:
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        if col > 3: 
            cell.border = Border(left=thick_side, bottom=thin_side, top=thin_side, right=thin_side)
        else:
            cell.border = Border(bottom=thin_side, top=thin_side, right=thin_side, left=thin_side)
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+span-1)
        col += span

    headers_det = [
        "S.N.", "Symbol", "Company",
        "Qty", "Rate", "Amt", "Qty", "Rate", "Amt", "Qty", "Rate", "Amt", 
        "Qty", "Rate", "Amt", "Consump", 
        "Real. P/L",
        "Qty", "WACC", "Cost", 
        ltp_header, "Value", "Unreal P/L", 
        "Total Profit", "Cash Div", "Total (incl. Div)"
    ]
    section_starts = {4, 7, 10, 13, 16, 17, 20, 23}

    for c_idx, title in enumerate(headers_det, 1):
        cell = ws.cell(row=2, column=c_idx, value=title)
        cell.font = font_header
        cell.alignment = align_center
        cell.fill = fill_header
        left_style = thick_side if c_idx in section_starts else thin_side
        cell.border = Border(left=left_style, bottom=thick_side, right=thin_side)

    # --- DATA ROWS ---
    current_row = 3
    
    for sector, content in data.items():
        # SECTOR HEADER (SUB TOTAL)
        ws.cell(row=current_row, column=2, value=sector).font = font_subtotal
        ws.cell(row=current_row, column=3, value="Sub Total").font = font_subtotal
        
        for c in range(1, 27):
            cell = ws.cell(row=current_row, column=c)
            cell.fill = fill_subtotal
            cell.font = font_subtotal
            left_s = thick_side if c in section_starts else None
            cell.border = Border(left=left_s, bottom=thin_side, top=thin_side)

        def write_sub(col, val, is_pl=False):
            c = ws.cell(row=current_row, column=col, value=val)
            c.number_format = num_fmt
            c.alignment = align_right
            left_s = thick_side if col in section_starts else None
            c.border = Border(left=left_s, bottom=thin_side, top=thin_side)
            if is_pl and val:
                if val < 0: 
                    c.font = Font(name='Calibri', size=9, bold=True, color="9C0006")
                elif val > 0: 
                    c.font = Font(name='Calibri', size=9, bold=True, color="006100")

        write_sub(4, content['totals']['op_kitta'])
        write_sub(6, content['totals']['op_amt'])
        write_sub(7, content['totals']['buy_kitta'])
        write_sub(9, content['totals']['buy_amt'])
        write_sub(10, content['totals']['bonus_kitta'])
        write_sub(12, content['totals']['bonus_amt'])
        write_sub(13, content['totals']['sale_kitta'])
        write_sub(15, content['totals']['sale_amt'])
        write_sub(16, content['totals']['consumption'])
        write_sub(17, content['totals']['realized_pl'], is_pl=True)
        write_sub(18, content['totals']['cl_kitta'])
        write_sub(20, content['totals']['cl_cost'])
        write_sub(22, content['totals']['market_val'])
        write_sub(23, content['totals']['unrealized_pl'], is_pl=True)
        write_sub(24, content['totals']['total_pl'], is_pl=True)
        write_sub(25, content['totals']['cash_dividend'])
        # Calculate total including dividend
        total_incl_div = content['totals']['total_pl'] + content['totals']['cash_dividend']
        write_sub(26, total_incl_div, is_pl=True)
        
        current_row += 1

        # DATA ROWS
        num_rows = len(content['rows'])
        if num_rows > 0:
            for r_idx in range(current_row, current_row + num_rows):
                ws.row_dimensions[r_idx].outlineLevel = 1

        for r in content['rows']:
            c1 = ws.cell(row=current_row, column=1, value=r['sn'])
            c1.alignment = align_center
            c1.font = font_body
            c1.border = Border(bottom=thin_side)
            
            c2 = ws.cell(row=current_row, column=2, value=r['company'])
            c2.alignment = align_left
            c2.font = Font(name='Calibri', size=9, bold=True)
            c2.border = Border(bottom=thin_side)
            
            c3 = ws.cell(row=current_row, column=3, value=r['company_name'])
            c3.alignment = align_left
            c3.font = font_body
            c3.border = Border(bottom=thin_side)

            def write_val(col, val, fmt=num_fmt, is_pl=False):
                c = ws.cell(row=current_row, column=col, value=val)
                c.font = font_body
                c.number_format = fmt
                c.alignment = align_right
                left_s = thick_side if col in section_starts else None
                c.border = Border(left=left_s, bottom=thin_side)
                if is_pl and val:
                    if val < 0: 
                        c.fill = fill_loss
                    elif val > 0: 
                        c.fill = fill_profit
            
            # Opening (columns 4-6)
            write_val(4, r['op_kitta'] or 0)
            write_val(5, r['op_rate'], dec_fmt)
            write_val(6, r['op_amt'])
            
            # Purchase (columns 7-9)
            write_val(7, r['buy_kitta'] or 0)
            write_val(8, r['buy_rate'], dec_fmt)
            write_val(9, r['buy_amt'])
            
            # Bonus (columns 10-12)
            write_val(10, r['bonus_kitta'] or 0)
            write_val(11, r['bonus_rate'], dec_fmt)
            write_val(12, r['bonus_amt'])
            
            # Sales (columns 13-16)
            write_val(13, r['sale_kitta'] or 0)
            write_val(14, r['sale_rate'], dec_fmt)
            write_val(15, r['sale_amt'])
            write_val(16, r['consumption'])
            
            # Performance (column 17 only)
            write_val(17, r['realized_pl'], is_pl=True)
            
            # Closing (columns 18-20)
            write_val(18, r['cl_kitta'])
            write_val(19, r['cl_rate'], dec_fmt)
            write_val(20, r['cl_cost'])
            
            # Market Valuation (columns 21-23)
            write_val(21, r['ltp'], dec_fmt)
            write_val(22, r['market_val'])
            write_val(23, r['unrealized_pl'], is_pl=True)
            
            # Net P/L (columns 24-26)
            write_val(24, r['total_pl'], is_pl=True)
            write_val(25, r['cash_dividend'])
            # Calculate total including dividend
            total_incl_div = r['total_pl'] + r['cash_dividend']
            write_val(26, total_incl_div, is_pl=True)
            
            current_row += 1

    # --- GRAND TOTAL ROW ---
    ws.cell(row=current_row, column=2, value="GRAND TOTAL").font = font_grand
    ws.cell(row=current_row, column=3, value="GRAND TOTAL").font = font_grand
    
    for c in range(1, 27):
        cell = ws.cell(row=current_row, column=c)
        cell.fill = fill_grand
        cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        if c not in [2, 3]: 
            cell.value = ""

    def write_grand(col, val, is_pl=False):
        c = ws.cell(row=current_row, column=col, value=val)
        c.font = font_grand
        c.number_format = num_fmt
        c.alignment = align_right
        c.fill = fill_grand
        left_s = Side(style='medium', color="FFFFFF") if col in section_starts else None
        c.border = Border(left=left_s, top=thick_side, bottom=thick_side)
        if is_pl and val:
            if val < 0: 
                c.font = Font(name='Calibri', size=10, bold=True, color="FF9999")
            elif val > 0: 
                c.font = Font(name='Calibri', size=10, bold=True, color="99FF99")

    write_grand(6, totals['op_amt'])
    write_grand(9, totals['buy_amt'])
    write_grand(12, totals['bonus_amt'])
    write_grand(15, totals['sale_amt'])
    write_grand(16, totals['consumption'])
    write_grand(17, totals['realized_pl'], True)
    write_grand(20, totals['cl_cost'])
    write_grand(22, totals['market_val'])
    write_grand(23, totals['unrealized_pl'], True)
    write_grand(24, totals['total_pl'], True)
    write_grand(25, totals['cash_dividend'])
    # Calculate grand total including dividend
    grand_total_incl_div = totals['total_pl'] + totals['cash_dividend']
    write_grand(26, grand_total_incl_div, True)

    # --- COLUMN WIDTHS ---
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    for c in range(4, 27):
        ws.column_dimensions[get_column_letter(c)].width = 12

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Valuation_Report_{end_date}.xlsx'
    wb.save(response)
    return response

# --- ADD THESE NEW VIEWS FOR BROKER TRANSACTIONS ---
@login_required
def broker_transaction_list_and_add(request):
    if request.method == 'POST':
        try:
            broker_no = request.POST.get('broker')
            date = request.POST.get('date')
            action = request.POST.get('action')
            amount_str = request.POST.get('amount')
            remarks = request.POST.get('remarks', '')

            if not broker_no or not date or not action or not amount_str:
                messages.error(request, "Error: Missing required fields (Broker, Date, Action, Amount).")
                return redirect('my_portfolio:broker_transactions')
            
            try:
                broker = Brokers.objects.get(broker_no=broker_no)
            except Brokers.DoesNotExist:
                messages.error(request, f"Error: Broker {broker_no} not found.")
                return redirect('my_portfolio:broker_transactions')
            
            try:
                amount = Decimal(amount_str)
            except InvalidOperation:
                messages.error(request, "Error: Invalid Amount format.")
                return redirect('my_portfolio:broker_transactions')

            new_txn = BrokerTransaction(
                broker=broker,
                date=date,
                action=action,
                amount=amount,
                remarks=remarks
            )
            new_txn.save() 
            messages.success(request, "Broker transaction added successfully!")
            
        except Exception as e:
            messages.error(request, f"An unexpected server error occurred: {str(e)}")
        
        return redirect('my_portfolio:broker_transactions')

    transactions_list = BrokerTransaction.objects.all().select_related('broker')
    filter_broker = request.GET.get('filter_broker', '')
    filter_action = request.GET.get('filter_action', '')
    
    if filter_broker:
        transactions_list = transactions_list.filter(broker__broker_no=filter_broker)
    if filter_action:
        transactions_list = transactions_list.filter(action=filter_action)

    rows_per_page = request.GET.get('rows', '20')
    
    if rows_per_page == 'all':
        page_obj = transactions_list
        is_paginated = False
    else:
        try:
            rows_int = int(rows_per_page)
        except ValueError:
            rows_int = 20
        
        paginator = Paginator(transactions_list, rows_int)
        page_num = request.GET.get('page', 1)
        is_paginated = True
        
        try:
            page_obj = paginator.page(page_num)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

    brokers = Brokers.objects.all().order_by('broker_no')
    action_choices = BrokerTransaction.ActionType.choices
    current_filters = {'broker': filter_broker, 'action': filter_action, 'rows': rows_per_page}
    filter_params = f"&filter_broker={filter_broker}&filter_action={filter_action}&rows={rows_per_page}"

    context = {
        'page_obj': page_obj,
        'is_paginated': is_paginated,
        'brokers': brokers,
        'action_choices': action_choices,
        'current_filters': current_filters,
        'filter_params': filter_params,
        'rows_options': ['20', '50', '100', 'all'],
    }
    return render(request, 'my_portfolio/broker_transactions.html', context)


@login_required
def broker_transaction_edit(request, unique_id):
    # (This view is correct and remains unchanged)
    txn = get_object_or_404(BrokerTransaction, unique_id=unique_id)
    if request.method == 'POST':
        try:
            broker_no = request.POST.get('broker')
            date = request.POST.get('date')
            action = request.POST.get('action')
            amount_str = request.POST.get('amount')
            remarks = request.POST.get('remarks', '')

            broker = Brokers.objects.get(broker_no=broker_no)
            amount = Decimal(amount_str)
            
            txn.broker = broker; txn.date = date; txn.action = action
            txn.amount = amount; txn.remarks = remarks
            txn.save()
            
            messages.success(request, "Transaction updated successfully.")
            return redirect('my_portfolio:broker_transactions')
        except Exception as e:
            messages.error(request, f"Error updating transaction: {e}")
            return redirect('my_portfolio:broker_transaction_edit', unique_id=unique_id)

    brokers = Brokers.objects.all().order_by('broker_no')
    context = {
        'transaction': txn,
        'brokers': brokers,
        'action_choices': BrokerTransaction.ActionType.choices
    }
    return render(request, 'my_portfolio/edit_broker_transaction.html', context)


@login_required
@require_POST
def broker_transaction_delete(request, unique_id):
    # (This view is correct and remains unchanged)
    txn = get_object_or_404(BrokerTransaction, unique_id=unique_id)
    try:
        txn.delete()
        messages.success(request, "Transaction deleted.")
    except Exception as e:
        messages.error(request, f"Error deleting transaction: {e}")
    return redirect('my_portfolio:broker_transactions')

@login_required
@require_POST
@db_transaction.atomic
def broker_transaction_upload(request):
    file = request.FILES.get('file')
    if not file or not file.name.endswith('.csv'):
        messages.error(request, "Please upload a valid CSV file.")
        return redirect('my_portfolio:broker_transactions')

    success_count = 0
    error_count = 0
    errors = []
    
    valid_broker_nos = set(Brokers.objects.values_list('broker_no', flat=True))
    valid_actions = set(BrokerTransaction.ActionType.values)

    try:
        csv_file = TextIOWrapper(file, encoding='utf-8', errors='replace')
        reader = csv.DictReader(csv_file)
        reader.fieldnames = [header.strip() for header in reader.fieldnames]
        
        required_headers = ['Date', 'Broker', 'Action', 'Amount']
        if not all(header in reader.fieldnames for header in required_headers):
            missing = [h for h in required_headers if h not in reader.fieldnames]
            messages.error(request, f"File missing required columns: {', '.join(missing)}")
            return redirect('my_portfolio:broker_transactions')

        for index, row in enumerate(reader, start=2):
            try:
                date_str = str(row.get('Date', '')).split()[0].strip()
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                broker_no_str = str(row.get('Broker', '')).strip()
                if not broker_no_str.isdigit():
                    raise ValueError(f"Broker '{broker_no_str}' must be a number.")
                broker_no = int(broker_no_str)
                if broker_no not in valid_broker_nos:
                    raise ValueError(f"Broker {broker_no} not found in database.")
                
                broker = Brokers.objects.get(broker_no=broker_no)
                
                action = str(row.get('Action', '')).strip()
                if action not in valid_actions:
                    raise ValueError(f"Invalid Action '{action}'. Must be one of: {', '.join(valid_actions)}")

                amount_str = str(row.get('Amount', '')).strip()
                if not amount_str:
                    raise ValueError("Amount cannot be empty.")
                amount = Decimal(amount_str)
                # Allowing negative amounts for cash ledger entries.

                remarks = str(row.get('Remarks', '')).strip() or None

                BrokerTransaction(
                    broker=broker,
                    date=date,
                    action=action,
                    amount=amount,
                    remarks=remarks
                ).save()
                success_count += 1

            except Exception as e:
                errors.append(f"Row {index}: {str(e)}")
                error_count += 1
                continue

        if error_count > 0:
            db_transaction.set_rollback(True)
            messages.error(request, f"Upload failed. {error_count} errors found. First error: {errors[0]}")
        else:
            messages.success(request, f"Upload successful! {success_count} broker transactions added.")

    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {e}")

    return redirect('my_portfolio:broker_transactions')


@login_required
def download_broker_template(request):
    """
    Provides a CSV template for broker R/P transactions.
    """
    fieldnames = ['Date', 'Broker', 'Action', 'Amount', 'Remarks']
    sample_data = [
        {'Date': '2025-11-14', 'Broker': '58', 'Action': 'Payment', 'Amount': '150000', 'Remarks': 'Fund transfer for buy'},
        {'Date': '2025-11-15', 'Broker': '45', 'Action': 'Receipt', 'Amount': '25000', 'Remarks': 'Sale proceeds'},
    ]

    output = TextIOWrapper(BytesIO(), encoding='utf-8', newline='')
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(sample_data)
    output.flush()

    response = HttpResponse(output.buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="broker_transaction_template.csv"'
    return response

def get_broker_rp_entries(broker_no):
    """Fetches and normalizes R/P entries for the ledger."""
    rp_entries = []
    
    # Define the sign convention for the cash ledger (positive = Debit/Inflow, negative = Credit/Outflow)
    # The convention can depend on your accounting standard, but we'll use a standard ledger view:
    # Debits (Broker pays you/Your balance increases): Receipt, Misc(+)
    # Credits (You pay broker/Your balance decreases): Payment, Chq Issue, Pledge Charge, Misc(-)
    
    debit_actions = {'Balance b/d', 'Receipt', 'Misc(+)'}
    credit_actions = {'Payment', 'Chq Issue', 'Pledge Charge', 'Misc(-)'}

    # Fetch all BrokerTransactions for the broker
    txns = BrokerTransaction.objects.filter(
        broker__broker_no=broker_no
    ).select_related('broker').order_by('date', 'created_at')

    for txn in txns:
        amount = txn.amount
        is_debit = False
        
        if txn.action in debit_actions:
            # Positive amounts for Debit actions remain Debit
            is_debit = True
        elif txn.action in credit_actions:
            # Positive amounts for Credit actions are registered as Credit
            is_debit = False
            
        # Handle the special case of Balance b/d, where the sign is already set by the user
        if txn.action == 'Balance b/d':
            is_debit = (amount >= 0)
        
        rp_entries.append({
            'date': txn.date,
            'description': f"{txn.get_action_display()} - {txn.remarks or ''}",
            'source': 'CASH',
            'amount': amount,
            'debit': amount if is_debit else Decimal('0.00'),
            'credit': abs(amount) if not is_debit else Decimal('0.00')
        })
        
    return rp_entries

def _get_broker_ledger_data(broker_no, sort='asc'):
    cash_txns = BrokerTransaction.objects.filter(
        broker__broker_no=broker_no
    ).order_by('date', 'created_at')

    ob_entries = cash_txns.filter(action='Balance b/d')
    other_cash_txns = cash_txns.exclude(action='Balance b/d')
    opening_balance = ob_entries.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.0')

    stock_txns = Transaction.objects.filter(
        broker=str(broker_no) 
    ).select_related('symbol').order_by('date', 'created_at')

    all_entries = []
    
    for txn in other_cash_txns:
        amount = txn.amount
        is_debit = False
        if txn.action in ['Receipt', 'Misc(+)']: is_debit = True
        all_entries.append({
            'date': txn.date,
            'description': f"{txn.get_action_display()} - {txn.remarks or ''}",
            'source': 'CASH',
            'debit': amount if is_debit else Decimal('0.00'),
            'credit': abs(amount) if not is_debit else Decimal('0.00')
        })

    for txn in stock_txns:
        amount = txn.billed_amount or Decimal('0.0')
        if txn.transaction_type in ['SALE', 'CONVERSION(-)', 'SUSPENSE(-)']:
            all_entries.append({
                'date': txn.date,
                'description': f"Stock {txn.transaction_type} of {txn.symbol.script_ticker} ({txn.kitta} kitta)",
                'source': 'STOCK', 'debit': amount, 'credit': Decimal('0.00')
            })
        elif txn.transaction_type in ['BUY', 'IPO', 'RIGHT', 'CONVERSION(+)', 'SUSPENSE(+)']:
            all_entries.append({
                'date': txn.date,
                'description': f"Stock {txn.transaction_type} of {txn.symbol.script_ticker} ({txn.kitta} kitta)",
                'source': 'STOCK', 'debit': Decimal('0.00'), 'credit': amount
            })
        elif txn.transaction_type == 'CASH':
             all_entries.append({
                'date': txn.date,
                'description': f"Cash Dividend for {txn.symbol.script_ticker}",
                'source': 'STOCK', 'debit': amount, 'credit': Decimal('0.00')
            })
            
    all_entries.sort(key=lambda x: x['date'], reverse=(sort == 'desc'))

    running_balance = opening_balance
    total_debit = Decimal('0.00')
    total_credit = Decimal('0.00')
    ledger = []

    for entry in all_entries:
        running_balance += entry['debit'] - entry['credit']
        total_debit += entry['debit']
        total_credit += entry['credit']
        entry['running_balance'] = running_balance
        ledger.append(entry)

    return {
        'opening_balance': opening_balance,
        'ledger': ledger,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'final_balance': running_balance
    }

@login_required
def broker_ledger_report(request):
    
    # 1. Get ONLY brokers who have transactions
    rp_brokers = set(BrokerTransaction.objects.values_list('broker__broker_no', flat=True).distinct())
    sp_brokers_str = set(Transaction.objects.values_list('broker', flat=True).distinct())
    sp_brokers = {int(b) for b in sp_brokers_str if b and b.isdigit()}
    
    active_broker_nos = rp_brokers.union(sp_brokers)
    all_brokers = Brokers.objects.filter(broker_no__in=active_broker_nos).order_by('broker_no')
    
    # 2. Get filter/pagination parameters
    selected_broker_no = request.GET.get('broker', '')
    current_sort = request.GET.get('sort', 'asc')
    current_rows = request.GET.get('rows', '50')
    rows_options = ['50', '100', '200', 'all']

    ledger_data = None
    broker = None
    page_obj = None
    is_paginated = False

    # 3. If a broker was selected, get their data
    if selected_broker_no:
        try:
            broker = get_object_or_404(Brokers, broker_no=selected_broker_no)
            
            # 4. Call the helper function to get all ledger data (unpaginated)
            ledger_data = _get_broker_ledger_data(selected_broker_no, current_sort)
            
            # 5. Paginate the results
            ledger_list = ledger_data['ledger']
            if current_rows != 'all':
                paginator = Paginator(ledger_list, int(current_rows))
                page_num = request.GET.get('page', 1)
                is_paginated = True
                try:
                    page_obj = paginator.page(page_num)
                except PageNotAnInteger:
                    page_obj = paginator.page(1)
                except EmptyPage:
                    page_obj = paginator.page(paginator.num_pages)
            else:
                page_obj = ledger_list # Not paginated

        except:
            messages.error(request, f"Broker {selected_broker_no} not found.")
            
    # 6. Create filter_params to preserve state in links
    filter_params = f"&broker={selected_broker_no}&sort={current_sort}&rows={current_rows}"
    
    # 7. Pass everything to the template
    context = {
        'all_brokers': all_brokers,
        'selected_broker_no': selected_broker_no,
        'broker': broker,
        'ledger_data': ledger_data,  # Contains totals
        'page_obj': page_obj,        # Contains the paginated list of entries
        'is_paginated': is_paginated,
        'current_sort': current_sort,
        'current_rows': current_rows,
        'rows_options': rows_options,
        'filter_params': filter_params # For pagination links
    }
    
    return render(request, 'my_portfolio/broker_ledger_report.html', context)


@login_required
def download_broker_ledger(request):
    broker_no = request.GET.get('broker', '')
    sort = request.GET.get('sort', 'asc')

    if not broker_no:
        messages.error(request, "No broker selected for download.")
        return redirect('my_portfolio:broker_ledger_report')

    try:
        broker = get_object_or_404(Brokers, broker_no=broker_no)
        
        # 1. Get all data (unpaginated)
        ledger_data = _get_broker_ledger_data(broker_no, sort)
        
        # 2. Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ledger_{broker_no}_{broker.name}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow(['Date', 'Source', 'Description', 'Debit (DR)', 'Credit (CR)', 'Balance'])
        
        # Write Opening Balance
        writer.writerow(['', '', 'Opening Balance', '', '', ledger_data['opening_balance']])
        
        # Write ledger rows
        for entry in ledger_data['ledger']:
            writer.writerow([
                entry['date'],
                entry['source'],
                entry['description'],
                entry['debit'],
                entry['credit'],
                entry['running_balance']
            ])
            
        # Write totals
        writer.writerow([])
        writer.writerow(['', '', 'Total (Excl. OB)', ledger_data['total_debit'], ledger_data['total_credit'], ''])
        writer.writerow(['', '', 'Final Balance', '', '', ledger_data['final_balance']])
        
        return response

    except Exception as e:
        messages.error(request, f"Error generating download: {e}")
        return redirect('my_portfolio:broker_ledger_report')
    

@login_required
def api_broker_settlement_summary(request):
    try:
        # 1. Get parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        sort_by = request.GET.get('sort_by', 'final_balance')
        sort_dir = request.GET.get('sort_dir', 'desc')
        
        # --- Default Date Logic (Find min/max dates) ---
        rp_dates = BrokerTransaction.objects.aggregate(min_date=Min('date'), max_date=Max('date'))
        sp_dates = Transaction.objects.aggregate(min_date=Min('date'), max_date=Max('date'))

        default_start_date = date.today()
        if rp_dates['min_date'] and sp_dates['min_date']:
            default_start_date = min(rp_dates['min_date'], sp_dates['min_date'])
        elif rp_dates['min_date']:
            default_start_date = rp_dates['min_date']
        elif sp_dates['min_date']:
            default_start_date = sp_dates['min_date']

        default_end_date = date.today()
        if rp_dates['max_date'] and sp_dates['max_date']:
            default_end_date = max(rp_dates['max_date'], sp_dates['max_date'])
        elif rp_dates['max_date']:
            default_end_date = rp_dates['max_date']
        elif sp_dates['max_date']:
            default_end_date = sp_dates['max_date']

        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start_date
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else default_end_date
        # --- End Date Logic ---
        
        # 2. Get all active brokers
        rp_brokers = set(BrokerTransaction.objects.values_list('broker__broker_no', flat=True).distinct())
        sp_brokers_str = set(Transaction.objects.values_list('broker', flat=True).distinct())
        sp_brokers = {int(b) for b in sp_brokers_str if b and b.isdigit()}
        active_broker_nos = rp_brokers.union(sp_brokers)
        
        brokers = Brokers.objects.filter(broker_no__in=active_broker_nos)
        
        # 3. Define transaction types
        CASH_DEBIT_ACTIONS = ['Receipt', 'Misc(+)']
        CASH_CREDIT_ACTIONS = ['Payment', 'Chq Issue', 'Pledge Charge', 'Misc(-)']
        STOCK_DEBIT_TYPES = ['SALE', 'CONVERSION(-)', 'SUSPENSE(-)']
        STOCK_CREDIT_TYPES = ['BUY', 'IPO', 'RIGHT', 'CONVERSION(+)', 'SUSPENSE(+)']

        # 4. Calculate for each broker
        summary_list = []
        for broker in brokers:
            broker_no = broker.broker_no
            broker_no_str = str(broker_no)

            # --- Opening Balance (all txns before start_date) ---
            cash_ob_txns = BrokerTransaction.objects.filter(broker__broker_no=broker_no, date__lt=start_date)
            cash_ob_balance_bd = cash_ob_txns.filter(action='Balance b/d').aggregate(total=Coalesce(Sum('amount'), Decimal(0)))['total']
            cash_ob_debit = cash_ob_txns.filter(action__in=CASH_DEBIT_ACTIONS).aggregate(total=Coalesce(Sum('amount'), Decimal(0)))['total']
            cash_ob_credit = cash_ob_txns.filter(action__in=CASH_CREDIT_ACTIONS).aggregate(total=Coalesce(Sum('amount'), Decimal(0)))['total']
            op_balance_cash = cash_ob_balance_bd + cash_ob_debit - cash_ob_credit
            
            stock_ob_debit = Transaction.objects.filter(
                broker=broker_no_str, transaction_type__in=STOCK_DEBIT_TYPES, date__lt=start_date
            ).aggregate(total=Coalesce(Sum('billed_amount'), Decimal(0)))['total']
            stock_ob_credit = Transaction.objects.filter(
                broker=broker_no_str, transaction_type__in=STOCK_CREDIT_TYPES, date__lt=start_date
            ).aggregate(total=Coalesce(Sum('billed_amount'), Decimal(0)))['total']
            op_balance_stock = stock_ob_debit - stock_ob_credit
            op_balance = op_balance_cash + op_balance_stock
            
            # --- Period Movements (between start_date and end_date) ---
            
            # A. Total Receipt (Cash In)
            total_cash_debit = BrokerTransaction.objects.filter(
                broker__broker_no=broker_no, date__range=[start_date, end_date], action__in=CASH_DEBIT_ACTIONS
            ).aggregate(total=Coalesce(Sum('amount'), Decimal(0)))['total']
            
            # B. Total Sale (Stock In)
            total_stock_debit = Transaction.objects.filter(
                broker=broker_no_str, transaction_type__in=STOCK_DEBIT_TYPES, date__range=[start_date, end_date]
            ).aggregate(total=Coalesce(Sum('billed_amount'), Decimal(0)))['total']

            # C. Total Payment (Cash Out)
            total_cash_credit = BrokerTransaction.objects.filter(
                broker__broker_no=broker_no, date__range=[start_date, end_date], action__in=CASH_CREDIT_ACTIONS
            ).aggregate(total=Coalesce(Sum('amount'), Decimal(0)))['total']
            
            # D. Total Buy (Stock Out)
            total_stock_credit = Transaction.objects.filter(
                broker=broker_no_str, transaction_type__in=STOCK_CREDIT_TYPES, date__range=[start_date, end_date]
            ).aggregate(total=Coalesce(Sum('billed_amount'), Decimal(0)))['total']
            
            # E. Final Balance
            final_balance = op_balance + (total_cash_debit + total_stock_debit) - (total_cash_credit + total_stock_credit)
            
            # --- UPDATED SUMMARY DICTIONARY ---
            summary_list.append({
                "broker_no": broker_no,
                "broker_name": broker.name,
                "op_balance": op_balance,
                "total_sale": total_stock_debit,    # Renamed
                "total_receipt": total_cash_debit,  # Renamed
                "total_buy": total_stock_credit,    # Renamed
                "total_payment": total_cash_credit, # Renamed
                "final_balance": final_balance,
            })
            
        # 5. Sort the list
        summary_list.sort(key=lambda x: x[sort_by], reverse=(sort_dir == 'desc'))
        
        # 6. Calculate Totals
        grand_total = {
            "op_balance": sum(item['op_balance'] for item in summary_list),
            "total_sale": sum(item['total_sale'] for item in summary_list),
            "total_receipt": sum(item['total_receipt'] for item in summary_list),
            "total_buy": sum(item['total_buy'] for item in summary_list),
            "total_payment": sum(item['total_payment'] for item in summary_list),
            "final_balance": sum(item['final_balance'] for item in summary_list),
        }

        return JsonResponse({
            "summary_list": summary_list,
            "grand_total": grand_total,
            "start_date": start_date.strftime('%Y-%m-%d'),
            "end_date": end_date.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    

@login_required
def sp_report(request):
    # 1. Date Range Logic
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    selected_sector = request.GET.get('sector', '') 
    
    if end_date_str: 
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else: 
        end_date = timezone.now().date()
        
    if start_date_str: 
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    else:
        start_date = end_date - timedelta(days=7)

    # 2. Fetch Latest Prices (LTP) & Date
    latest_prices = {}
    ltp_date_str = ""
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """)
            rows = dictfetchall(cursor)
            if rows:
                valid_dates = [r['business_date'] for r in rows if r['business_date']]
                if valid_dates:
                    max_date = max(valid_dates)
                    ltp_date_str = max_date.strftime('%Y-%m-%d')
            for row in rows:
                price = Decimal(row['close_price']) if row['close_price'] else Decimal('0.0')
                latest_prices[row['symbol']] = price     
    except Exception as e:
        print(f"Error fetching prices: {e}")

    # [cite_start]3. Fetch Bonus Adjustments (The Fix) [cite: 3]
    # We only care about BONUS for Opp Cost because it's "free" extra quantity.
    # Rights require payment, which complicates the 'Cost' part of the equation.
    adjustments_map = defaultdict(list)
    try:
        # Filter only 'bonus' types
        adjustments = PriceAdjustments.objects.filter(adjustment_type__iexact='bonus')
        for adj in adjustments:
            # Store relevant data: Date of book close and the %
            adjustments_map[adj.symbol.script_ticker].append({
                'date': adj.book_close_date,
                'percent': adj.adjustment_percent
            })
    except Exception as e:
        print(f"Error fetching adjustments: {e}")

    # 4. Fetch Transactions
    all_transactions = Transaction.objects.all().select_related('symbol').order_by('symbol__script_ticker', 'date', 'created_at')
    
    grouped_txns = defaultdict(list)
    available_sectors = set() 

    for txn in all_transactions:
        if txn.symbol.sector:
            available_sectors.add(txn.symbol.sector)
        txn_dict = {
            'unique_id': txn.unique_id,
            'date': txn.date,
            'broker': txn.broker,
            'transaction_type': txn.transaction_type,
            'kitta': txn.kitta,
            'billed_amount': txn.billed_amount,
            'eff_rate': txn.eff_rate,
            'symbol_ticker': txn.symbol.script_ticker,
            'sector': txn.symbol.sector
        }
        grouped_txns[txn.symbol.script_ticker].append(txn_dict)

    buy_agg = defaultdict(lambda: {'p_kitta': 0, 't_purchase': Decimal('0.0'), 'ltp': Decimal('0.0'), 'sector': '', 'simulated_kitta': Decimal('0.0')})
    sell_agg = defaultdict(lambda: {'s_kitta': 0, 'total_sales': Decimal('0.0'), 'profit': Decimal('0.0'), 'ltp': Decimal('0.0'), 'sector': '', 'simulated_kitta': Decimal('0.0')})
    
    # 5. Process each symbol
    for symbol, txns in grouped_txns.items():
        company_sector = txns[0].get('sector', '')
        if selected_sector and selected_sector != 'All' and company_sector != selected_sector:
            continue

        ltp = latest_prices.get(symbol, Decimal('0.0'))
        
        dummy_price_info = {'close_price': ltp, 'business_date': None}
        detailed_calculations, _ = calculate_pma_details(txns, dummy_price_info)
        
        # Get adjustments for this specific symbol
        symbol_adjustments = adjustments_map.get(symbol, [])
        
        for row in detailed_calculations:
            if start_date <= row['date'] <= end_date:
                
                # --- CALCULATE MULTIPLIER ---
                # Check if any bonus happened AFTER this transaction
                qty_multiplier = Decimal('1.0')
                for adj in symbol_adjustments:
                    if adj['date'] > row['date']:
                        # Formula: Multiplier * (1 + percent/100)
                        # E.g., 20% bonus -> 1 * 1.2 = 1.2
                        factor = Decimal('1.0') + (adj['percent'] / Decimal('100.0'))
                        qty_multiplier *= factor

                if row['is_buy'] and row['type'] not in ('Balance b/d', 'CASH'):
                    p_kitta = row['p_qty']
                    p_amount = row['p_amount']
                    
                    if p_kitta > 0:
                        buy_agg[symbol]['p_kitta'] += p_kitta
                        buy_agg[symbol]['t_purchase'] += p_amount
                        buy_agg[symbol]['ltp'] = ltp
                        buy_agg[symbol]['sector'] = company_sector
                        # Add the "Adjusted" Kitta for Opp Cost calculation
                        buy_agg[symbol]['simulated_kitta'] += (Decimal(p_kitta) * qty_multiplier)

                elif row['is_sale']:
                    s_kitta = row['s_qty']
                    s_amount = row['s_amount']
                    profit = row['profit']
                    
                    if s_kitta > 0:
                        sell_agg[symbol]['s_kitta'] += s_kitta
                        sell_agg[symbol]['total_sales'] += s_amount
                        sell_agg[symbol]['profit'] += profit
                        sell_agg[symbol]['ltp'] = ltp
                        sell_agg[symbol]['sector'] = company_sector
                        # Add the "Adjusted" Kitta for Opp Cost calculation
                        sell_agg[symbol]['simulated_kitta'] += (Decimal(s_kitta) * qty_multiplier)

    # 6. Aggregate Lists & Calculate Adjusted Prices
    buy_data = []
    buy_totals = {'kitta': 0, 'amount': Decimal('0.0'), 'opp_cost': Decimal('0.0')}
    
    for symbol, data in buy_agg.items():
        kitta = data['p_kitta']
        amount = data['t_purchase']
        ltp = data['ltp']
        sim_kitta = data['simulated_kitta'] 
        
        avg_rate = (amount / Decimal(kitta)) if kitta > 0 else Decimal('0.0')
        opp_cost = (ltp * sim_kitta) - amount
        
        # Calculate Effective Adjusted LTP for Display
        # This shows what the price "feels like" relative to the original quantity
        adj_ltp = ltp
        if kitta > 0 and sim_kitta != kitta:
            adj_ltp = (ltp * sim_kitta) / Decimal(kitta)
        
        buy_data.append({
            'symbol': symbol,
            'sector': data['sector'],
            'p_kitta': kitta,
            't_purchase': amount,
            'rate': avg_rate,
            'ltp': ltp,
            'adj_ltp': adj_ltp, # New Field
            'opp_cost': opp_cost
        })
        
        buy_totals['kitta'] += kitta
        buy_totals['amount'] += amount
        buy_totals['opp_cost'] += opp_cost

    sell_data = []
    sell_totals = {'kitta': 0, 'amount': Decimal('0.0'), 'profit': Decimal('0.0'), 'opp_cost': Decimal('0.0')}
    
    for symbol, data in sell_agg.items():
        kitta = data['s_kitta']
        amount = data['total_sales']
        profit = data['profit']
        ltp = data['ltp']
        sim_kitta = data['simulated_kitta'] 
        
        avg_rate = (amount / Decimal(kitta)) if kitta > 0 else Decimal('0.0')
        opp_cost = (ltp * sim_kitta) - amount

        # Calculate Effective Adjusted LTP for Display
        adj_ltp = ltp
        if kitta > 0 and sim_kitta != kitta:
            adj_ltp = (ltp * sim_kitta) / Decimal(kitta)
        
        sell_data.append({
            'symbol': symbol,
            'sector': data['sector'],
            's_kitta': kitta,
            'total_sales': amount,
            'rate': avg_rate,
            'profit': profit,
            'ltp': ltp,
            'adj_ltp': adj_ltp, # New Field
            'opp_cost': opp_cost
        })
        
        sell_totals['kitta'] += kitta
        sell_totals['amount'] += amount
        sell_totals['profit'] += profit
        sell_totals['opp_cost'] += opp_cost

    buy_data.sort(key=lambda x: x['t_purchase'], reverse=True)
    sell_data.sort(key=lambda x: x['total_sales'], reverse=True)
    sorted_sectors = sorted(list(available_sectors))

    context = {
        'buy_data': buy_data,
        'buy_totals': buy_totals,
        'sell_data': sell_data,
        'sell_totals': sell_totals,
        'start_date': start_date,
        'end_date': end_date,
        'ltp_date': ltp_date_str,
        'available_sectors': sorted_sectors,
        'selected_sector': selected_sector
    }
    return render(request, 'my_portfolio/sp_report.html', context)

def portfolio_watch(request):
    # --- 1. Get Current NEPSE Index ---
    current_index_val = Decimal('0.0')
    latest_index_date = None
    
    try:
        index_obj = Indices.objects.filter(sector__icontains='NEPSE').order_by('-date').first()
        if index_obj:
            current_index_val = index_obj.close
            latest_index_date = index_obj.date
    except Exception as e:
        print(f"Error fetching index: {e}")
        current_index_val = Decimal('2700.00')

    # --- 2. Handle User Input ---
    expected_index_input = request.GET.get('expected_index')
    if expected_index_input:
        try:
            expected_index_val = Decimal(expected_index_input)
        except:
            expected_index_val = current_index_val
    else:
        expected_index_val = current_index_val

    # --- 3. Calculate Market Change % ---
    market_change_pct = Decimal('0.0')
    if current_index_val > 0:
        market_change_pct = (expected_index_val - current_index_val) / current_index_val

    # --- 4. Fetch Metadata Dates ---
    beta_updated_date = None
    portfolio_updated_date = None # <--- NEW VARIABLE
    
    try:
        # Beta Date
        beta_updated_date = CompanyMetrics.objects.filter(metric_item='Beta Weekly').aggregate(Max('business_date'))['business_date__max']
        
        # Portfolio Updated Date (Last Transaction Date) <--- NEW LOGIC
        last_txn = Transaction.objects.order_by('-date').first()
        if last_txn:
            portfolio_updated_date = last_txn.date
        else:
            portfolio_updated_date = date.today()
            
    except:
        pass

    # --- 5. Fetch Portfolio Data ---
    try:
        latest_prices = {}
        with connection.cursor() as cursor:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """)
            for row in dictfetchall(cursor):
                latest_prices[row['symbol']] = {
                    'close_price': row.get('close_price') or Decimal('0.0'),
                    'business_date': row.get('business_date')
                }

        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM my_portfolio_transaction ORDER BY symbol_id, date, created_at")
            all_transactions = dictfetchall(cursor)

        _, holdings_summary_list = calculate_overall_portfolio(all_transactions, latest_prices)
        held_symbols = [h['symbol'] for h in holdings_summary_list]
        fundamental_data = get_fundamental_metrics(held_symbols)

    except Exception as e:
        messages.error(request, f"Error loading data: {e}")
        holdings_summary_list = []
        fundamental_data = {}

    # --- 6. Process Data ---
    holdings_summary_list.sort(key=lambda x: (x.get('sector', ''), x['symbol']))

    sector_data = defaultdict(lambda: {
        'rows': [], 'total_cost': Decimal('0.0'), 'total_value': Decimal('0.0'), 
        'total_forecast_value': Decimal('0.0'), 'total_max_loss_1w': Decimal('0.0'), 'total_max_loss_1m': Decimal('0.0')
    })
    
    grand_total = {
        'cost': Decimal('0.0'), 'value': Decimal('0.0'), 'forecast_value': Decimal('0.0'),
        'max_loss_1w': Decimal('0.0'), 'max_loss_1m': Decimal('0.0')
    }

    sn_counter = 1
    for h in holdings_summary_list:
        symbol = h['symbol']
        metrics = fundamental_data.get(symbol, {})
        
        beta_val = metrics.get('beta')
        beta = Decimal(str(beta_val)) if beta_val is not None else Decimal('1.0')
        
        stock_change_pct = beta * market_change_pct
        ltp = h['ltp']
        forecast_price = ltp * (Decimal('1.0') + stock_change_pct)
        
        closing_qty = h['closing_kitta']
        market_val = ltp * closing_qty
        forecast_val = forecast_price * closing_qty

        var_1w = Decimal(str(metrics.get('var_1w') or 0))
        var_1m = Decimal(str(metrics.get('var_1m') or 0))
        max_loss_1w = market_val * (var_1w / Decimal('100.0'))
        max_loss_1m = market_val * (var_1m / Decimal('100.0'))

        row = {
            'sn': sn_counter, 'symbol': symbol, 'script': h['script'],
            'kitta': closing_qty, 'wacc': h['wacc'], 'ltp': ltp, 'market_val': market_val,
            'beta': beta, 'forecast_price': forecast_price, 'forecast_val': forecast_val,
            'forecast_diff': forecast_val - market_val, 'forecast_pct_change': stock_change_pct * 100,
            'var_1w': var_1w, 'max_loss_1w': max_loss_1w, 'var_1m': var_1m, 'max_loss_1m': max_loss_1m
        }
        sn_counter += 1
        sec = h['sector']
        sector_data[sec]['rows'].append(row)
        
        sector_data[sec]['total_cost'] += h['book_value']
        sector_data[sec]['total_value'] += market_val
        sector_data[sec]['total_forecast_value'] += forecast_val
        sector_data[sec]['total_max_loss_1w'] += max_loss_1w
        sector_data[sec]['total_max_loss_1m'] += max_loss_1m

        grand_total['cost'] += h['book_value']
        grand_total['value'] += market_val
        grand_total['forecast_value'] += forecast_val
        grand_total['max_loss_1w'] += max_loss_1w
        grand_total['max_loss_1m'] += max_loss_1m

    for sec, data in sector_data.items():
        data['total_forecast_diff'] = data['total_forecast_value'] - data['total_value']

    grand_total['forecast_diff'] = grand_total['forecast_value'] - grand_total['value']
    sorted_sector_data = dict(sorted(sector_data.items()))

    context = {
        'current_index': current_index_val,
        'latest_index_date': latest_index_date,
        'beta_updated_date': beta_updated_date,
        'portfolio_updated_date': portfolio_updated_date, # <--- Passed to template
        'expected_index': expected_index_val,
        'market_change_pct': market_change_pct * 100,
        'sector_data': sorted_sector_data,
        'grand_total': grand_total
    }
    return render(request, 'my_portfolio/portfolio_watch.html', context)


@login_required
def download_portfolio_watch(request):
    # --- 1. Re-run Logic to Get Data ---
    # (We duplicate logic here to ensure download matches view exactly)
    try:
        current_index_val = Decimal('0.0')
        index_obj = Indices.objects.filter(sector__icontains='NEPSE').order_by('-date').first()
        if index_obj: current_index_val = index_obj.close
        else: current_index_val = Decimal('2700.00')

        expected_index_input = request.GET.get('expected_index')
        expected_index_val = Decimal(expected_index_input) if expected_index_input else current_index_val

        market_change_pct = Decimal('0.0')
        if current_index_val > 0:
            market_change_pct = (expected_index_val - current_index_val) / current_index_val

        latest_prices = {}
        with connection.cursor() as cursor:
            cursor.execute("""
                WITH RankedPrices AS (
                    SELECT symbol, close_price, business_date,
                        ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY business_date DESC) as rn
                    FROM stock_prices
                )
                SELECT symbol, close_price, business_date FROM RankedPrices WHERE rn = 1;
            """)
            for row in dictfetchall(cursor):
                latest_prices[row['symbol']] = {'close_price': row.get('close_price') or Decimal('0.0')}

        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM my_portfolio_transaction ORDER BY symbol_id, date, created_at")
            all_transactions = dictfetchall(cursor)

        _, holdings_summary_list = calculate_overall_portfolio(all_transactions, latest_prices)
        held_symbols = [h['symbol'] for h in holdings_summary_list]
        fundamental_data = get_fundamental_metrics(held_symbols)
        
        holdings_summary_list.sort(key=lambda x: (x.get('sector', ''), x['symbol']))

    except Exception as e:
        return HttpResponse(f"Error generating data: {e}")

    # --- 2. Create Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portfolio Watch"
    
    # Styles
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal='center')
    align_right = Alignment(horizontal='right')
    border_thin = Side(style='thin')
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    fill_header = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Info Header
    ws['A1'] = f"Portfolio Watch Report (Expected NEPSE: {expected_index_val})"
    ws['A1'].font = Font(size=12, bold=True)
    
    headers = [
        "S.N.", "Symbol", "Sector", "Kitta", "WACC", "LTP", "Market Value",
        "Beta", "Fcst Price", "Fcst Value", "Fcst Diff",
        "VaR 1W %", "Max Loss 1W", "VaR 1M %", "Max Loss 1M"
    ]
    
    ws.append([]) # Spacer
    ws.append(headers)
    
    # Header Style
    for cell in ws[3]:
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    sn = 1
    row_num = 4
    
    for h in holdings_summary_list:
        symbol = h['symbol']
        metrics = fundamental_data.get(symbol, {})
        beta_val = metrics.get('beta')
        beta = Decimal(str(beta_val)) if beta_val is not None else Decimal('1.0')
        
        stock_change_pct = beta * market_change_pct
        ltp = h['ltp']
        forecast_price = ltp * (Decimal('1.0') + stock_change_pct)
        closing_qty = h['closing_kitta']
        market_val = ltp * closing_qty
        forecast_val = forecast_price * closing_qty
        
        var_1w = Decimal(str(metrics.get('var_1w') or 0))
        var_1m = Decimal(str(metrics.get('var_1m') or 0))
        max_loss_1w = market_val * (var_1w / Decimal('100.0'))
        max_loss_1m = market_val * (var_1m / Decimal('100.0'))

        row_data = [
            sn, symbol, h['sector'], closing_qty, h['wacc'], ltp, market_val,
            beta, forecast_price, forecast_val, forecast_val - market_val,
            var_1w, max_loss_1w, var_1m, max_loss_1m
        ]
        ws.append(row_data)
        sn += 1
        row_num += 1

    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="portfolio_watch.xlsx"'
    wb.save(response)
    return response

@login_required
@db_transaction.atomic
def my_share_details(request):
    
    # 1. Handle POST
    if request.method == 'POST':
        
        # ... (add_demat logic is unchanged) ...
        if 'add_demat' in request.POST:
            # Copy your existing logic here
            pass 

        # --- UPLOAD MY SHARE (With Date) ---
        elif 'upload_csv' in request.POST:
            upload_form = MeroShareUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                demat_account = upload_form.cleaned_data['demat_account']
                target_date = upload_form.cleaned_data['snapshot_date'] # <--- GET DATE
                file = request.FILES['file']
                
                try:
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    
                    def clean_decimal(val):
                        if not val: return Decimal(0)
                        return Decimal(str(val).replace(',', '').strip())
                    
                    success_count = 0
                    for row in reader:
                        scrip = row.get('Scrip', '').strip().upper()
                        if not scrip: continue
                        
                        try:
                            company = Companies.objects.get(script_ticker=scrip)
                        except Companies.DoesNotExist:
                            continue

                        # Update or Create using snapshot_date
                        MeroShareHolding.objects.update_or_create(
                            demat_account=demat_account,
                            symbol=company,
                            snapshot_date=target_date, # <--- USE DATE
                            defaults={
                                'current_balance': clean_decimal(row.get('Current Balance')),
                                'pledge_balance': clean_decimal(row.get('Pledge Balance')),
                                'lockin_balance': clean_decimal(row.get('Lockin Balance')),
                                'freeze_balance': clean_decimal(row.get('Freeze Balance')),
                                'free_balance': clean_decimal(row.get('Free Balance')),
                                'demat_pending': clean_decimal(row.get('Demat Pending')),
                                'remarks': row.get('Remarks', '')
                            }
                        )
                        success_count += 1
                    
                    messages.success(request, f"Synced {success_count} scripts for date {target_date}!")
                    # Redirect with the date in URL so user sees what they just uploaded
                    return redirect(f"{reverse('my_portfolio:my_share_details')}?filter_date={target_date}")
                    
                except Exception as e:
                    messages.error(request, f"Error: {e}")

        # --- UPLOAD EDIS (With Date) ---
        elif 'upload_transfer_csv' in request.POST:
            transfer_form = TransferRequestUploadForm(request.POST, request.FILES)
            if transfer_form.is_valid():
                demat_account = transfer_form.cleaned_data['demat_account']
                target_date = transfer_form.cleaned_data['settlement_date'] # <--- GET DATE
                file = request.FILES['file']
                
                try:
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    TARGET_STATUSES = {'PENDING', 'ACKNOWLEDGED'}
                    
                    processed_count = 0
                    modified_ids = []
                    for row in reader:
                        status = row.get('Status', '').strip().upper()
                        if status not in TARGET_STATUSES: continue
                        
                        scrip = row.get('Scrip', '').strip().upper()
                        try:
                            qty_to_transfer = Decimal(row.get('Quantity', '0').replace(',', ''))
                        except: continue
                        if qty_to_transfer <= 0: continue

                        try:
                            # ONLY find holdings for the SPECIFIC DATE
                            holding = MeroShareHolding.objects.get(
                                demat_account=demat_account, 
                                symbol__script_ticker=scrip,
                                snapshot_date=target_date # <--- FILTER BY DATE
                            )
                        except MeroShareHolding.DoesNotExist:
                            continue

                        # ... (Calculations same as before) ...
                        current_free = holding.free_balance
                        current_pledge = holding.pledge_balance
                        current_total = holding.current_balance
                        
                        if qty_to_transfer >= current_total:
                            holding.current_balance = Decimal(0)
                            holding.free_balance = Decimal(0)
                            holding.pledge_balance = Decimal(0)
                            # ... zero others ...
                        else:
                            holding.current_balance -= qty_to_transfer
                            if qty_to_transfer <= current_free:
                                holding.free_balance -= qty_to_transfer
                            else:
                                remainder = qty_to_transfer - current_free
                                holding.free_balance = Decimal(0)
                                if remainder > current_pledge:
                                    holding.pledge_balance = Decimal(0)
                                else:
                                    holding.pledge_balance -= remainder
                        
                        holding.save()
                        modified_ids.append(holding.id)
                        processed_count += 1
                    request.session['highlight_ids'] = modified_ids

                    messages.success(request, f"Applied {processed_count} transfers to records of {target_date}.")
                    return redirect(f"{reverse('my_portfolio:my_share_details')}?filter_date={target_date}")

                except Exception as e:
                    messages.error(request, f"Error: {e}")

        # --- D. Delete Holdings ---
        elif 'delete_holdings' in request.POST:
            delete_date_str = request.POST.get('delete_date')
            delete_demat_id = request.POST.get('delete_demat_account')

            if delete_date_str and delete_demat_id:
                try:
                    target_date = datetime.strptime(delete_date_str, '%Y-%m-%d').date()
                    filters = Q(updated_at__date=target_date)
                    
                    if delete_demat_id != 'ALL':
                        filters &= Q(demat_account_id=delete_demat_id)
                    
                    deleted_count, _ = MeroShareHolding.objects.filter(filters).delete()
                    
                    if deleted_count > 0:
                        messages.success(request, f"Deleted {deleted_count} records dated {target_date}.")
                    else:
                        messages.warning(request, f"No records found for date {target_date}.")
                        
                except ValueError:
                    messages.error(request, "Invalid Date format.")
                except Exception as e:
                    messages.error(request, f"Error deleting records: {e}")
            else:
                messages.error(request, "Please select both a Date and a DP Provider.")
            
            return redirect('my_portfolio:my_share_details')

    # 2. Prepare Context (GET)
    demat_accounts = DematAccount.objects.all()
    
    # --- DATE FILTERING LOGIC ---
    # By default, show TODAY or LATEST available date
    filter_date_str = request.GET.get('filter_date')
    selected_dp_id = request.GET.get('dp_id', 'ALL')
    
    holdings_query = MeroShareHolding.objects.select_related('symbol', 'demat_account').all()

    # 1. Filter by DP
    if selected_dp_id and selected_dp_id != 'ALL' and selected_dp_id.isdigit():
        holdings_query = holdings_query.filter(demat_account_id=int(selected_dp_id))

    # 2. Determine Date
    available_dates = MeroShareHolding.objects.values_list('snapshot_date', flat=True).distinct().order_by('-snapshot_date')
    
    if filter_date_str:
        try:
            display_date = datetime.strptime(filter_date_str, '%Y-%m-%d').date()
        except:
            display_date = date.today()
    else:
        # Default to latest available date, or today if empty
        display_date = available_dates.first() if available_dates else date.today()

    # 3. Apply Date Filter
    holdings = holdings_query.filter(snapshot_date=display_date)
    highlight_ids = request.session.pop('highlight_ids', [])

    context = {
        'demat_form': DematAccountForm(),
        'upload_form': MeroShareUploadForm(),
        'transfer_form': TransferRequestUploadForm(),
        'demat_accounts': demat_accounts,
        'holdings': holdings,
        'display_date': display_date, # To show in UI
        'available_dates': available_dates, # For dropdown
        'selected_dp_id': selected_dp_id,
        'highlight_ids': highlight_ids, # <--- Pass to template
    }
    
    return render(request, 'my_portfolio/my_share_details.html', context)



@login_required
def download_my_share_csv(request):
    """
    Downloads the current view of the holdings table based on the selected filter.
    """
    # Get the same filter from URL
    dp_id = request.GET.get('dp_id', 'ALL')
    
    # Base Query
    holdings = MeroShareHolding.objects.select_related('symbol', 'demat_account').all()
    
    # Apply Filter
    filename_suffix = "All_DPs"
    if dp_id and dp_id != 'ALL' and dp_id.isdigit():
        holdings = holdings.filter(demat_account_id=int(dp_id))
        # Try to get provider name for filename
        try:
            provider = DematAccount.objects.get(id=int(dp_id))
            filename_suffix = provider.capital_name.replace(" ", "_")
        except:
            pass
            
    # Create Response
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    response['Content-Disposition'] = f'attachment; filename="Net_Holdings_{filename_suffix}_{timestamp}.csv"'
    
    writer = csv.writer(response)
    
    # Write Header
    writer.writerow([
        'Scrip', 'Current Balance', 'Free Balance', 'Pledge Balance', 
        'Lockin Balance', 'Freeze Balance', 'DP Provider', 'Last Updated'
    ])
    
    # Write Rows
    for h in holdings:
        writer.writerow([
            h.symbol.script_ticker,
            h.current_balance,
            h.free_balance,
            h.pledge_balance,
            h.lockin_balance,
            h.freeze_balance,
            h.demat_account.capital_name,
            h.updated_at.strftime('%Y-%m-%d %H:%M') if h.updated_at else ""
        ])
        
    return response

@login_required
def generate_trading_sheet(request):
    """
    Generates a Trading Sheet (PDF & Excel).
    - Values in CRORES.
    - Excel: Auto-fit Columns, Fit Sheet on One Page (1x1), No Margins.
    - Orientation: Portrait.
    - Alignment: Center (Data), Left (Script/Sector).
    """
    
    # --- A. PREPARE DATA ---
    
    # 1. Determine Report Date (Latest Snapshot)
    # Always check for the latest snapshot date first
    latest_snapshot = MeroShareHolding.objects.aggregate(Max('snapshot_date'))['snapshot_date__max']
    report_date = latest_snapshot if latest_snapshot else date.today()
    
    today_calc = date.today()
    start_date = date(2000, 1, 1) 
    
    # 2. Get Valuation Data
    raw_valuation_data, grand_totals = _get_valuation_data(start_date, today_calc)
    
    CRORE = Decimal("10000000")
    grand_bv_cr = grand_totals['cl_cost'] / CRORE
    grand_mv_cr = grand_totals['market_val'] / CRORE
    grand_vpl_cr = grand_totals['total_pl'] / CRORE

    # 3. Get Demat Providers
    demat_accounts = DematAccount.objects.all().order_by('id')
    demat_headers = []
    for acc in demat_accounts:
        name_upper = acc.capital_name.upper()
        if "GLOBAL" in name_upper: short_name = "GBIME"
        elif "IMPERIAL" in name_upper: short_name = "IMP"
        elif "NABIL" in name_upper: short_name = "NABIL"
        else: short_name = name_upper.split()[0][:5]
        demat_headers.append({'id': acc.id, 'name': short_name})

    # 4. Get Free Balances
    free_balance_map = defaultdict(dict)
    ms_holdings = MeroShareHolding.objects.all()
    for h in ms_holdings:
        free_balance_map[h.symbol.script_ticker][h.demat_account.id] = h.free_balance

    # 5. Get Technical Data
    technical_map = {}
    all_symbols = []
    for sector_data in raw_valuation_data.values():
        for row in sector_data['rows']:
            if row['cl_kitta'] > 0:
                all_symbols.append(row['company'])

    for symbol in all_symbols:
        price_obj = StockPrices.objects.filter(symbol=symbol).order_by('-business_date').first()
        if price_obj:
            high = float(price_obj.high_price or 0)
            low = float(price_obj.low_price or 0)
            close = float(price_obj.close_price or 0)
            if high > 0:
                pp = (high + low + close) / 3
                technical_map[symbol] = {
                    's1': Decimal((2 * pp) - high),
                    's2': Decimal(pp - (high - low)),
                    's3': Decimal(low - 2 * (high - pp)),
                    'r1': Decimal((2 * pp) - low),
                    'r2': Decimal(pp + (high - low)),
                    'r3': Decimal(high + 2 * (pp - low))
                }
            else: technical_map[symbol] = {}
        else: technical_map[symbol] = {}

    # 6. Enrich Portfolio Data
    sector_grouped_data = {}
    sn_counter = 1
    
    for sector in sorted(raw_valuation_data.keys()):
        content = raw_valuation_data[sector]
        enriched_rows = []
        
        sec_total_bv = Decimal(0)
        sec_total_mv = Decimal(0)
        sec_total_vpl = Decimal(0)
        
        for row in content['rows']:
            if row['cl_kitta'] <= 0: continue
                
            symbol = row['company']
            qty = row['cl_kitta']
            bv = row['cl_cost']
            wacc = bv / qty if qty > 0 else 0
            ltp = row['ltp']
            mv = qty * ltp
            vpl = mv - bv
            pl_pct = (vpl / bv * 100) if bv > 0 else 0
            
            sec_total_bv += bv
            sec_total_mv += mv
            sec_total_vpl += vpl

            free_bals = []
            for d in demat_headers:
                bal = free_balance_map.get(symbol, {}).get(d['id'], 0)
                free_bals.append(bal)

            tech = technical_map.get(symbol, {})
            
            item = {
                'sn': sn_counter,
                'grade': '', 
                'script': symbol,
                'qty': qty,
                'wacc': wacc,
                'bv_cr': bv / CRORE,
                'ltp': ltp,
                'mv_cr': mv / CRORE,
                'vpl_cr': vpl / CRORE,
                'pl_pct': pl_pct,
                'free_balances': free_bals,
                'target_buy': tech.get('s1', 0),
                'price_sale': tech.get('r1', 0),
                's2': tech.get('s2', 0),
                'r2': tech.get('r2', 0),
            }
            enriched_rows.append(item)
            sn_counter += 1
        
        if enriched_rows:
            pct_of_bv = (sec_total_bv / grand_totals['cl_cost'] * 100) if grand_totals['cl_cost'] > 0 else 0
            pct_of_mv = (sec_total_mv / grand_totals['market_val'] * 100) if grand_totals['market_val'] > 0 else 0
            sec_pl_pct = (sec_total_vpl / sec_total_bv * 100) if sec_total_bv > 0 else 0
            
            sector_grouped_data[sector] = {
                'rows': enriched_rows,
                'totals': {
                    'bv_str': f"{(sec_total_bv/CRORE):,.2f} ({pct_of_bv:.0f}%)",
                    'mv_str': f"{(sec_total_mv/CRORE):,.2f} ({pct_of_mv:.0f}%)",
                    'vpl_cr': sec_total_vpl / CRORE,
                    'pl_pct_str': f"{sec_pl_pct:.1f}%"
                }
            }

    # --- B. GENERATE EXCEL ---
    if request.GET.get('type') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = "Trading Sheet"
        
        # Styles
        style_title = Font(size=14, bold=True, color="000000") 
        style_header = Font(bold=True, color="FFFFFF")
        fill_header = PatternFill("solid", fgColor="000000") 
        fill_super = PatternFill("solid", fgColor="808080") 
        fill_sector = PatternFill("solid", fgColor="D9D9D9")
        fill_free_bal = PatternFill("solid", fgColor="FDE9D9") # Orange Lighter 80%
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # Title with Snapshot Date
        ws.merge_cells('A1:F1'); ws['A1'] = "TRADING SHEET"; ws['A1'].font = style_title; ws['A1'].alignment = center_align
        ws.merge_cells('A2:F2'); ws['A2'] = f"Date: {report_date.strftime('%Y-%m-%d')}"; ws['A2'].font = Font(bold=True); ws['A2'].alignment = center_align

        # Columns
        col_free_start = 11
        col_free_end = 10 + len(demat_headers)
        col_tech_start = col_free_end + 1
        col_tech_end = col_tech_start + 3

        # Super Headers
        ws.merge_cells(start_row=3, start_column=col_free_start, end_row=3, end_column=col_free_end)
        c = ws.cell(row=3, column=col_free_start, value="Free Balance")
        c.font = style_header; c.fill = fill_super; c.alignment = center_align

        ws.merge_cells(start_row=3, start_column=col_tech_start, end_row=3, end_column=col_tech_end)
        c = ws.cell(row=3, column=col_tech_start, value="Support/Resistance")
        c.font = style_header; c.fill = fill_super; c.alignment = center_align

        # Main Headers
        headers = ["SN", "Grade", "Script", "Qty", "Rate", "BV (Cr)", "LTP", "MV (Cr)", "V P/L (Cr)", "P/L %"]
        for d in demat_headers: headers.append(d['name'])
        headers.extend(["Tgt Buy", "Prc Sale", "S2", "R2"])
        
        ws.append([]) 
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = fill_header; cell.font = style_header; cell.alignment = center_align; cell.border = thin_border

        # Data Rows
        for sector, data in sector_grouped_data.items():
            rows = data['rows']; totals = data['totals']

            # Sector Header with Totals
            ws.append([sector.upper(), "", "", "", "", 
                       totals['bv_str'], "", totals['mv_str'], float(totals['vpl_cr']), totals['pl_pct_str']])
            sec_row_idx = ws.max_row
            ws.merge_cells(start_row=sec_row_idx, start_column=1, end_row=sec_row_idx, end_column=5)
            
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=sec_row_idx, column=c)
                cell.font = Font(bold=True); cell.fill = fill_sector; cell.border = thin_border
                if c == 1: cell.alignment = left_align # Sector Name Left
                else: cell.alignment = center_align    # Totals Center
                if c == 9: cell.number_format = '#,##0.00'

            # Stock Rows
            for r in rows:
                row_data = [
                    r['sn'], r['grade'], r['script'], float(r['qty']), float(r['wacc']), 
                    float(r['bv_cr']), float(r['ltp']), float(r['mv_cr']), float(r['vpl_cr']), 
                    f"{r['pl_pct']:.2f}%"
                ]
                for bal in r['free_balances']: row_data.append(float(bal) if bal > 0 else 0)
                row_data.extend([float(r['target_buy']), float(r['price_sale']), float(r['s2']), float(r['r2'])])
                ws.append(row_data)
                
                curr_row = ws.max_row
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=curr_row, column=col)
                    cell.border = thin_border
                    
                    # Highlight Free Balance
                    if col_free_start <= col <= col_free_end:
                        cell.font = Font(bold=True)
                        cell.fill = fill_free_bal 

                    if col == 4 or (11 <= col < col_tech_start): cell.number_format = '#,##0'
                    elif col in [5, 6, 7, 8, 9] or col >= col_tech_start: cell.number_format = '#,##0.00'
                    
                    # Alignment
                    if col == 3: cell.alignment = left_align # Script Left
                    else: cell.alignment = center_align      # Others Center
                    
                    if col == 10:
                        cell.font = Font(color="FF0000" if r['pl_pct'] < 0 else "006100")

        # Grand Total
        ws.append(["GRAND TOTAL", "", "", "", "", 
                   f"{grand_bv_cr:,.2f}", "", f"{grand_mv_cr:,.2f}", f"{grand_vpl_cr:,.2f}", ""])
        grand_row = ws.max_row
        ws.merge_cells(start_row=grand_row, start_column=1, end_row=grand_row, end_column=5)
        
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=grand_row, column=c)
            cell.fill = PatternFill("solid", fgColor="000000")
            cell.font = Font(bold=True, color="FFFFFF")
            if c == 1: cell.alignment = left_align
            else: cell.alignment = center_align

        # --- PRINT SETTINGS (FIT SHEET ON ONE PAGE) ---
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 1  # Force 1 Page High
        ws.page_setup.fitToWidth = 1   # Force 1 Page Wide
        
        # Remove Margins for Printing
        ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1, header=0.0, footer=0.0)
        ws.print_options.horizontalCentered = True

        # Auto-fit Columns
        for i, col_cells in enumerate(ws.columns, 1):
            max_len = 0; col_letter = get_column_letter(i)
            for cell in col_cells:
                if cell.row <= 3: continue 
                try:
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_len + 2.5, 50)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Trading_Sheet_{report_date}.xlsx"'
        wb.save(response)
        return response

    # --- C. GENERATE PDF ---
    else:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Trading_Sheet_{report_date}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2, leftMargin=2, topMargin=2, bottomMargin=2)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TinyTitle', parent=styles['Title'], fontSize=10, alignment=1)
        
        elements.append(Paragraph(f"<b>TRADING SHEET</b>", title_style))
        elements.append(Paragraph(f"As of: {report_date.strftime('%Y-%m-%d')}", ParagraphStyle('Sub', parent=styles['Normal'], fontSize=6, alignment=1)))
        elements.append(Spacer(1, 4))

        base_headers = ["SN", "Gr", "Script", "Qty", "Rate", "BV(Cr)", "MR", "MV(Cr)", "VPL(Cr)", "P/L%"]
        demat_names = [d['name'] for d in demat_headers]
        tech_headers = ["Buy", "Sale", "S2", "R2"]
        all_headers = base_headers + demat_names + tech_headers
        total_cols = len(all_headers)

        idx_script = 2 
        idx_free_start = 10 
        idx_free_end = 10 + len(demat_names)

        weights = [0.5, 0.4, 1.1, 0.9, 0.7, 1.0, 0.7, 1.0, 0.9, 0.8] + ([0.8]*len(demat_names)) + [0.7, 0.7, 0.7, 0.7]
        unit_w = 591 / sum(weights)
        col_widths = [w * unit_w for w in weights]

        table_data = [all_headers]
        
        tbl_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'), # Default Center
            ('ALIGN', (idx_script, 0), (idx_script, -1), 'LEFT'), # Script Left
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 5),
            ('GRID', (0, 0), (-1, -1), 0.1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 4.5),
            ('TOPPADDING', (0, 0), (-1, -1), 0.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0.5),
            ('FONTNAME', (idx_free_start, 0), (idx_free_end-1, 0), 'Helvetica-BoldOblique'),
        ]

        row_idx = 1
        for sector, data in sector_grouped_data.items():
            rows = data['rows']; totals = data['totals']
            
            # Sector Row
            sec_row = [sector.upper(), '', '', '', '', 
                       totals['bv_str'], '', totals['mv_str'], f"{totals['vpl_cr']:,.2f}", totals['pl_pct_str']]
            sec_row += [''] * (len(demat_names) + 4)
            table_data.append(sec_row)
            
            tbl_style.append(('SPAN', (0, row_idx), (4, row_idx)))
            tbl_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey))
            tbl_style.append(('ALIGN', (0, row_idx), (0, row_idx), 'LEFT')) # Sector Left
            tbl_style.append(('ALIGN', (5, row_idx), (-1, row_idx), 'CENTER')) # Totals Center
            tbl_style.append(('FONTSIZE', (0, row_idx), (-1, row_idx), 5))
            tbl_style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            tbl_style.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.black))
            row_idx += 1

            for r in rows:
                def fmt_num(n): return f"{n:,.0f}"
                def fmt_cr(n): return f"{n:,.2f}"
                row = [
                    r['sn'], '', r['script'], fmt_num(r['qty']), fmt_num(r['wacc']), 
                    fmt_cr(r['bv_cr']), fmt_num(r['ltp']), fmt_cr(r['mv_cr']), 
                    fmt_cr(r['vpl_cr']), f"{r['pl_pct']:.1f}%"
                ]
                for bal in r['free_balances']: row.append(fmt_num(bal) if bal > 0 else "-")
                row.extend([fmt_num(r['target_buy']), fmt_num(r['price_sale']), fmt_num(r['s2']), fmt_num(r['r2'])])
                table_data.append(row)
                
                color = colors.red if r['pl_pct'] < 0 else colors.green
                tbl_style.append(('TEXTCOLOR', (9, row_idx), (9, row_idx), color))
                
                # Highlight Free Balance
                tbl_style.append(('BACKGROUND', (idx_free_start, row_idx), (idx_free_end-1, row_idx), colors.HexColor('#FDE9D9')))
                tbl_style.append(('FONTNAME', (idx_free_start, row_idx), (idx_free_end-1, row_idx), 'Helvetica-Bold'))
                row_idx += 1

        grand_row = ['GRAND TOTAL', '', '', '', '', 
                     f"{grand_bv_cr:,.2f}", '', f"{grand_mv_cr:,.2f}", f"{grand_vpl_cr:,.2f}", ""]
        grand_row += [''] * (len(demat_names) + 4)
        table_data.append(grand_row)
        tbl_style.append(('SPAN', (0, row_idx), (4, row_idx)))
        tbl_style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
        tbl_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.black))
        tbl_style.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.white))
        tbl_style.append(('ALIGN', (0, row_idx), (0, row_idx), 'LEFT'))

        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        main_table.setStyle(TableStyle(tbl_style))
        elements.append(main_table)

        doc.build(elements)
        return response